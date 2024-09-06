import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# Functie om een Excel-bestand te maken met een tabel
def maak_excel():
    # Ophalen van de waarden uit de inputvelden
    bestandsnaam = bestandsnaam_entry.get()
    tabelnaam = tabelnaam_entry.get()
    aantal_kolommen = int(kolommen_entry.get())
    aantal_werkbladen = int(werkbladen_entry.get())

    # Controleren of de invoer geldig is
    if not bestandsnaam or not tabelnaam:
        messagebox.showerror("Fout", "Bestandsnaam en tabelnaam zijn verplicht.")
        return

    # Bestandsnaam met .xlsx-extensie
    bestandsnaam = bestandsnaam + ".xlsx"

    # Excel-bestand en werkbladen aanmaken
    workbook = openpyxl.Workbook()

    # Aanmaken van werkbladen met tabellen
    for i in range(aantal_werkbladen):
        if i == 0:
            sheet = workbook.active
            sheet.title = f"{tabelnaam}_{i+1}"
        else:
            sheet = workbook.create_sheet(title=f"{tabelnaam}_{i+1}")
        
        # Kolomnamen toevoegen
        kolomnamen = []
        for col in range(1, aantal_kolommen + 1):
            kolom_naam = f"K {col}"
            sheet.cell(row=1, column=col).value = kolom_naam
            kolomnamen.append(kolom_naam)

        # Tabel aanmaken in het werkblad
        table_range = f"A1:{chr(64 + aantal_kolommen)}2"  # Een eenvoudige tabel met de opgegeven kolommen
        tabel = Table(displayName=f"{tabelnaam}_{i+1}", ref=table_range)

        # Stijl toepassen op de tabel
        stijl = TableStyleInfo(
            name="TableStyleLight8", 
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tabel.tableStyleInfo = stijl

        # Voeg de tabel toe aan het werkblad
        sheet.add_table(tabel)

    # Excel-bestand opslaan
    workbook.save(bestandsnaam)

    # Bevestiging en bestand openen
    messagebox.showinfo("Succes", f"Bestand {bestandsnaam} aangemaakt!")
    os.system(f'start excel "{bestandsnaam}"')

# GUI venster instellen
root = tk.Tk()
root.title("Excel Generator")

# Labels en inputvelden
tk.Label(root, text="Naam bestand:").grid(row=0)
bestandsnaam_entry = tk.Entry(root)
bestandsnaam_entry.grid(row=0, column=1)

tk.Label(root, text="Naam Tabel:").grid(row=1)
tabelnaam_entry = tk.Entry(root)
tabelnaam_entry.grid(row=1, column=1)

tk.Label(root, text="Hoeveel kolommen?").grid(row=2)
kolommen_entry = tk.Entry(root)
kolommen_entry.grid(row=2, column=1)

tk.Label(root, text="Hoeveel werkbladen?").grid(row=3)
werkbladen_entry = tk.Entry(root)
werkbladen_entry.grid(row=3, column=1)

# Knop om Excel-bestand te genereren
generate_button = tk.Button(root, text="Maak Excel", command=maak_excel)
generate_button.grid(row=4, column=1)

# GUI starten
root.mainloop()
