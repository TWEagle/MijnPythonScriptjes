import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo  # Import voor het maken van een tabel

def custom_style():
    style = {
        "bg": "black",
        "fg": "white",
        "font": ("Arial", 14),
        "button": {"bg": "dark green", "fg": "white", "activebackground": "green", "font": ("Arial", 14)}
    }
    return style

def browse_file():
    global input_file_path
    input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if input_file_path:
        file_label.config(text=os.path.basename(input_file_path))

def browse_folder():
    global output_dir
    output_dir = filedialog.askdirectory()
    if output_dir:
        folder_label.config(text=output_dir)

def add_table_to_sheet(worksheet, num_columns, table_name):
    """Voeg een Excel-tabel toe aan een werkblad."""
    # Bepaal het bereik voor de tabel
    table_range = f"A1:{get_column_letter(num_columns)}{worksheet.max_row}"
    
    # Maak een tabelobject aan
    table = Table(displayName=table_name, ref=table_range)

    # Definieer een stijl voor de tabel
    style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    
    # Voeg de tabel toe aan het werkblad
    worksheet.add_table(table)

def run_script():
    if not (input_file_path and output_dir):
        messagebox.showerror("Error", "Gelieve een bestand en map te selecteren.")
        return

    try:
        # Maak de submappen 'Gemeente' en 'Dienstenleverancier' aan
        gemeente_dir = os.path.join(output_dir, "Gemeente")
        dienstenleverancier_dir = os.path.join(output_dir, "Dienstenleverancier")
        os.makedirs(gemeente_dir, exist_ok=True)
        os.makedirs(dienstenleverancier_dir, exist_ok=True)

        # Lees de drie werkbladen in
        domeinen_df = pd.read_excel(input_file_path, sheet_name='Domeinen')
        toepassingen_df = pd.read_excel(input_file_path, sheet_name='Toepassingen')
        certificaten_df = pd.read_excel(input_file_path, sheet_name='Certificaten')

        # Zorg ervoor dat 'Nieuwe-Orgnaam' in alle werkbladen zit
        if 'Nieuwe-Orgnaam' not in domeinen_df.columns or 'Nieuwe-Orgnaam' not in toepassingen_df.columns or 'Nieuwe-Orgnaam' not in certificaten_df.columns:
            messagebox.showerror("Error", "De kolom 'Nieuwe-Orgnaam' ontbreekt in één of meerdere werkbladen.")
            return

        # Haal unieke 'Nieuwe-Orgnaam' en 'Nieuwe-Orgcode' op
        unieke_orgnamen = domeinen_df['Nieuwe-Orgnaam'].unique()

        # Aantal bestanden tellen
        gemeente_count = 0
        dienstenleverancier_count = 0

        # Loop door elke unieke 'Nieuwe-Orgnaam'
        for orgnaam in unieke_orgnamen:
            # Filter per werkblad op de huidige 'Nieuwe-Orgnaam'
            domeinen_filtered = domeinen_df[domeinen_df['Nieuwe-Orgnaam'] == orgnaam]
            toepassingen_filtered = toepassingen_df[toepassingen_df['Nieuwe-Orgnaam'] == orgnaam]
            certificaten_filtered = certificaten_df[certificaten_df['Nieuwe-Orgnaam'] == orgnaam]

            # Zorg ervoor dat 'Nieuwe-Orgcode' bestaat
            if 'Nieuwe-Orgcode' not in domeinen_filtered.columns:
                messagebox.showerror("Error", f"De kolom 'Nieuwe-Orgcode' ontbreekt in het werkblad 'Domeinen' voor orgnaam {orgnaam}.")
                continue

            orgcode = domeinen_filtered['Nieuwe-Orgcode'].iloc[0]  # Haal de 'Nieuwe-Orgcode' op
            output_path = os.path.join(gemeente_dir, f"{orgcode}.xlsx")  # Zet in de 'Gemeente' submap

            # Schrijf naar een nieuw Excel-bestand in de 'Gemeente' submap (overschrijven indien bestaat)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                domeinen_filtered.to_excel(writer, sheet_name='Domeinen', index=False)
                toepassingen_filtered.to_excel(writer, sheet_name='Toepassingen', index=False)
                certificaten_filtered.to_excel(writer, sheet_name='Certificaten', index=False)

                # Werk opmaak bij (zoals kolombreedte en vetgedrukte koppen en voeg de tabel toe)
                for sheet_name, num_columns, table_name in [('Domeinen', 5, 'Domeinen'),
                                                            ('Toepassingen', 8, 'Toepassingen'),
                                                            ('Certificaten', 8, 'Certificaten')]:
                    worksheet = writer.sheets[sheet_name]
                    for column_cells in worksheet.columns:
                        length = max(len(as_text(cell.value)) for cell in column_cells) + 2
                        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
                    for cell in worksheet["1:1"]:
                        cell.font = Font(bold=True, color="FFFFFF")  # Zet de headertekst naar wit
                    
                    # Voeg een tabel toe aan het werkblad
                    add_table_to_sheet(worksheet, num_columns, table_name)

            # Verhoog het aantal gemeente-bestanden
            gemeente_count += 1

        # Loop door elke unieke 'Delegatie-Orgcode' waar 'Delegatie' gelijk is aan 1
        for delegatie_orgcode in toepassingen_df[toepassingen_df['Delegatie'] == 1]['Delegatie-Orgcode'].unique():
            # Filter per werkblad op de huidige 'Delegatie-Orgcode' en waar 'Delegatie' gelijk is aan 1
            toepassingen_filtered = toepassingen_df[(toepassingen_df['Delegatie'] == 1) & (toepassingen_df['Delegatie-Orgcode'] == delegatie_orgcode)]
            certificaten_filtered = certificaten_df[(certificaten_df['Delegatie'] == 1) & (certificaten_df['Delegatie-Orgcode'] == delegatie_orgcode)]

            # Zorg ervoor dat 'Delegatie-Orgcode' bestaat en niet leeg is
            if delegatie_orgcode is None or delegatie_orgcode == '':
                messagebox.showerror("Error", f"De kolom 'Delegatie-Orgcode' ontbreekt voor een record in het werkblad.")
                continue

            delegatie_output_path = os.path.join(dienstenleverancier_dir, f"{delegatie_orgcode}.xlsx")  # Zet in de 'Dienstenleverancier' submap

            # Schrijf naar een nieuw Excel-bestand in de 'Dienstenleverancier' submap (overschrijven indien bestaat)
            with pd.ExcelWriter(delegatie_output_path, engine='openpyxl') as writer:
                if not toepassingen_filtered.empty:
                    toepassingen_filtered.to_excel(writer, sheet_name='Toepassingen', index=False)
                if not certificaten_filtered.empty:
                    certificaten_filtered.to_excel(writer, sheet_name='Certificaten', index=False)

                # Werk opmaak bij (zoals kolombreedte en vetgedrukte koppen en voeg de tabel toe)
                for sheet_name, num_columns, table_name in [('Toepassingen', 8, 'Toepassingen'),
                                                            ('Certificaten', 8, 'Certificaten')]:
                    worksheet = writer.sheets[sheet_name]
                    for column_cells in worksheet.columns:
                        length = max(len(as_text(cell.value)) for cell in column_cells) + 2
                        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
                    for cell in worksheet["1:1"]:
                        cell.font = Font(bold=True, color="FFFFFF")  # Zet de headertekst naar wit
                    
                    # Voeg een tabel toe aan het werkblad
                    add_table_to_sheet(worksheet, num_columns, table_name)

            # Verhoog het aantal dienstenleverancier-bestanden
            dienstenleverancier_count += 1

        # Toon een bericht met het aantal gegenereerde bestanden
        messagebox.showinfo("Voltooid", f"Proces is perfect doorlopen!\nGemeente-bestanden: {gemeente_count}\nDienstenleverancier-bestanden: {dienstenleverancier_count}")

        os.startfile(output_dir)

    except Exception as e:
        messagebox.showerror("Error", f"Er is een fout opgetreden: {str(e)}")

def as_text(value):
    if value is None:
        return ""
    return str(value)

def close_application():
    root.destroy()

root = tk.Tk()
root.title("Gemeentefusies Hulp Tool voor Excel")
style = custom_style()

root.config(bg=style["bg"])
root.geometry('1000x200')  # Adjust window size to make sure all elements are visible

# Create grid layout
file_button = tk.Button(root, text="Selecteer het Excel bestand", command=browse_file, **style["button"])
file_label = tk.Label(root, text="Geen bestand geselecteerd", bg=style["bg"], fg=style["fg"], font=style["font"])
folder_button = tk.Button(root, text="Selecteer de map waar alles opgeslagen wordt", command=browse_folder, **style["button"])
folder_label = tk.Label(root, text="Geen map geselecteerd", bg=style["bg"], fg=style["fg"], font=style["font"])
start_button = tk.Button(root, text="Starten", command=run_script, **style["button"])
close_button = tk.Button(root, text="Sluiten", command=close_application, **style["button"])

# Position with grid
file_button.grid(row=0, column=0, padx=(10, 20), pady=(10, 0), sticky='w')
file_label.grid(row=0, column=1, padx=(10, 20), pady=(10, 0), sticky='w')
folder_button.grid(row=1, column=0, padx=(10, 20), pady=(10, 0), sticky='w')
folder_label.grid(row=1, column=1, padx=(10, 20), pady=(10, 0), sticky='w')
start_button.grid(row=3, column=0, padx=(10, 20), pady=(10, 0), sticky='w')
close_button.grid(row=3, column=1, padx=(10, 20), pady=(10, 0), sticky='w')

root.mainloop()
