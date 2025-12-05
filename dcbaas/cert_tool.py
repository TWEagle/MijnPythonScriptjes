#!/usr/bin/env python3
"""
CyNiT Certificate / CSR Decoder Tool

Modes:
  - GUI + Web (default): python cert_tool.py
  - Alleen Web:          python cert_tool.py --web
  - Alleen GUI:          python cert_tool.py --gui-only
"""

# ============================================================
#   STANDARD LIB IMPORTS & GLOBAL FLAGS
# ============================================================

import sys
import os
import json
import csv
import threading
import subprocess
import webbrowser
import copy
from pathlib import Path
from io import BytesIO

# Flag: draaien we als PyInstaller EXE?
IS_FROZEN = bool(getattr(sys, "frozen", False))

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config" / "settings.json"

# ============================================================
#   THIRD PARTY IMPORTS
# ============================================================

# ---- Crypto imports ----
from cryptography import x509
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import rsa, dsa, ec
from cryptography.x509.oid import NameOID

# ---- XLSX export ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ---- Tkinter GUI imports ----
import tkinter as tk
from tkinter import (
    filedialog,
    messagebox,
    scrolledtext,  # niet meer gebruikt voor de tabel, maar laten staan kan geen kwaad
)

# ---- Flask web imports ----
from flask import (
    Flask,
    request,
    render_template_string,
    send_file,
    make_response,
    redirect,
    url_for,
)

# ---- Pillow (voor logo/ico manipulatie) ----
from PIL import Image, ImageTk


# ============================================================
#   DEFAULT ABOUT CONTENT (Markdown)
# ============================================================

ABOUT_DEFAULT = """# CyNiT Certificate / CSR Decoder

Deze tool is gemaakt om X.509 certificaten en CSRs snel te kunnen analyseren.

## Functionaliteit

- Ondersteuning voor certificaten en CSRs in PEM en DER
- GUI én Web interface
- Overzicht van:
  - Certificate Subject
  - Certificate Issuer
  - Certificate Properties
- Export naar:
  - JSON
  - CSV
  - XLSX
  - HTML
  - Markdown
- Automatische PyInstaller build van een standalone EXE
- CyNiT look & feel met thema-profielen via config/settings.json

## Gebruik

1. Start de tool:
   - `python cert_tool.py` voor GUI + Web
   - `python cert_tool.py --gui-only` alleen GUI
   - `python cert_tool.py --web` alleen Web

2. Kies een certificaat of CSR in de GUI of upload via de Web UI.

3. Bekijk de details of exporteer ze in het gewenste formaat.

## Profielen

- Alle kleuren, paden en UI-instellingen zitten in:
  - `config/settings.json`
- Je kunt meerdere profielen definiëren en wisselen:
  - in de GUI via de dropdown
  - in de webversie via /profiles

## Over

CyNiT tools zijn ontworpen om het beheer van certificaten en
security workflows eenvoudiger en leuker te maken.
"""


# ============================================================
#   CONFIG / PROFILE SYSTEM
# ============================================================

def default_config_dict():
    """Standaard config-structuur met 1 profiel: 'default'."""
    return {
        "active_profile": "default",
        "profiles": {
            "default": {
                "colors": {
                    "background": "#000000",
                    "general_fg": "#00B7C3",
                    "title": "#00A2FF",

                    "table_col1_bg": "#FEF102",
                    "table_col1_fg": "#000000",

                    "table_col2_bg": "#111111",
                    "table_col2_fg": "#00B7C3",

                    "button_bg": "#000000",
                    "button_fg": "#00B7C3"
                },
                "paths": {
                    "logo": "images/CyNiT-Logo.png",
                    "help": "ABOUT.md"
                },
                "ui": {
                    "logo_max_height": 80,
                    "font_main": "Consolas",
                    "font_buttons": "Segoe UI"
                }
            }
        }
    }


def deep_merge(default: dict, override: dict) -> dict:
    """Recursieve merge: override heeft voorrang, default vult aan."""
    result = dict(default)
    for k, v in override.items():
        if isinstance(v, dict) and k in result and isinstance(result[k], dict):
            result[k] = deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def save_config(cfg: dict):
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")


def load_or_init_config() -> dict:
    """Laadt config/settings.json, maakt en/of vult aan met defaults indien nodig."""
    default_cfg = default_config_dict()
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)

    if CONFIG_PATH.exists():
        try:
            raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            # corrupte config -> backup en reset
            backup = CONFIG_PATH.with_suffix(".bak")
            try:
                CONFIG_PATH.rename(backup)
            except Exception:
                pass
            raw = default_cfg
    else:
        raw = default_cfg
        save_config(raw)

    # Zorg dat basis-structuur bestaat
    if "profiles" not in raw or not isinstance(raw["profiles"], dict):
        raw["profiles"] = default_cfg["profiles"]

    if "active_profile" not in raw:
        raw["active_profile"] = "default"

    # Merge defaults in elke profile
    default_profile = default_cfg["profiles"]["default"]
    merged_profiles = {}
    for name, prof in raw["profiles"].items():
        if not isinstance(prof, dict):
            prof = {}
        merged_profiles[name] = deep_merge(default_profile, prof)

    raw["profiles"] = merged_profiles

    # active_profile valideren
    if raw["active_profile"] not in raw["profiles"]:
        raw["active_profile"] = "default"

    # Config eventueel terugschrijven (na opschoning)
    save_config(raw)
    return raw


def get_active_profile():
    cfg = load_or_init_config()
    name = cfg["active_profile"]
    profile = cfg["profiles"][name]
    return cfg, name, profile


CONFIG, ACTIVE_PROFILE_NAME, PROFILE = get_active_profile()

COLORS = PROFILE["colors"]
PATHS = PROFILE["paths"]
UI_CFG = PROFILE["ui"]

# Kleuren uit config
BG_COLOR = COLORS["background"]
MAIN_COLOR = COLORS["general_fg"]
TITLE_COLOR = COLORS["title"]

TABLE_COL1_BG = COLORS["table_col1_bg"]
TABLE_COL1_FG = COLORS["table_col1_fg"]

TABLE_COL2_BG = COLORS["table_col2_bg"]
TABLE_COL2_FG = COLORS["table_col2_fg"]

BUTTON_BG = COLORS["button_bg"]
BUTTON_FG = COLORS["button_fg"]

# Excel kleuren (zonder '#')
MAIN_COLOR_XLSX = MAIN_COLOR.lstrip("#")
TITLE_COLOR_XLSX = TITLE_COLOR.lstrip("#")

# Paden uit config
LOGO_BASE = BASE_DIR / PATHS["logo"]
ABOUT_MD = BASE_DIR / PATHS["help"]


def ensure_about_file():
    """Maakt een ABOUT.md met default inhoud als die nog niet bestaat."""
    if not ABOUT_MD.exists():
        try:
            ABOUT_MD.write_text(ABOUT_DEFAULT, encoding="utf-8")
        except Exception:
            pass

def restart_program():
    """
    Herstart het huidige Python/EXE proces met dezelfde argumenten.
    We gebruiken subprocess.Popen i.p.v. os.execv zodat paden met spaties
    zoals 'C:\\Program Files\\...' netjes werken onder Windows.
    """
    python = sys.executable
    try:
        # Start een nieuw proces met dezelfde arguments
        subprocess.Popen([python] + sys.argv, cwd=BASE_DIR)
    except Exception as e:
        # Als het misgaat, melden we het en stoppen gewoon
        print(f"[ERROR] Kon herstart niet uitvoeren: {e}")
    # Huidig proces hard afsluiten zodat de nieuwe netjes kan overnemen
    os._exit(0)

# ============================================================
#   WEB SERVER CONFIG
# ============================================================

WEB_HOST = "127.0.0.1"
WEB_PORT = 5000

_web_thread = None
_web_running = False

# Laatste gedecodeerde cert (alleen voor web-downloads)
LAST_INFO = None


def start_web_in_background():
    """Start de Flask webserver in een background thread (voor GUI+Web)."""
    global _web_thread, _web_running
    if _web_running:
        return

    def run():
        global _web_running
        _web_running = True
        try:
            app.run(host=WEB_HOST, port=WEB_PORT, debug=False, use_reloader=False)
        finally:
            _web_running = False

    _web_thread = threading.Thread(target=run, daemon=True)
    _web_thread.start()


# ============================================================
#   ICON HELPERS (PNG -> ICO VOOR FAVICON & PYINSTALLER)
# ============================================================

def _load_logo_image():
    """Laadt het master-logo als Pillow Image of None."""
    if not LOGO_BASE.exists():
        return None
    try:
        return Image.open(LOGO_BASE).convert("RGBA")
    except Exception:
        return None


def _make_square_icon_image():
    """Maakt een vierkante ico-afbeelding (Pillow Image) vanuit LOGO_BASE."""
    img = _load_logo_image()
    if img is None:
        return None

    size = 256
    square = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    scale = min(size / img.width, size / img.height)
    new_w = int(img.width * scale)
    new_h = int(img.height * scale)
    resized = img.resize((new_w, new_h), Image.LANCZOS)
    offset = ((size - new_w) // 2, (size - new_h) // 2)
    square.paste(resized, offset, resized)
    return square


def generate_ico_bytes():
    """
    Genereert ICO bytes in-memory vanuit het master-logo.
    Returnt None bij fout.
    """
    square = _make_square_icon_image()
    if square is None:
        return None

    ico_sizes = [(256, 256), (128, 128), (64, 64), (32, 32), (16, 16)]
    buf = BytesIO()
    try:
        square.save(buf, format="ICO", sizes=ico_sizes)
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


# ============================================================
#   CORE: X.509 / CSR DECODER
# ============================================================

def load_cert_or_csr(data: bytes):
    """
    Probeert PEM/DER voor Certificate of CSR.
    Return:
      ("cert", x509.Certificate) of ("csr", x509.CertificateSigningRequest)
    Raise:
      ValueError bij mislukking.
    """
    text = None
    try:
        text = data.decode("utf-8", errors="ignore")
    except Exception:
        pass

    # PEM CSR
    if text and "BEGIN CERTIFICATE REQUEST" in text:
        try:
            csr = x509.load_pem_x509_csr(data)
            return "csr", csr
        except Exception:
            pass

    # PEM CERT
    if text and "BEGIN CERTIFICATE" in text:
        try:
            cert = x509.load_pem_x509_certificate(data)
            return "cert", cert
        except Exception:
            pass

    # DER CERT
    try:
        cert = x509.load_der_x509_certificate(data)
        return "cert", cert
    except Exception:
        pass

    # DER CSR
    try:
        csr = x509.load_der_x509_csr(data)
        return "csr", csr
    except Exception:
        pass

    raise ValueError("Bestand is geen geldige X.509 certificate of CSR (PEM/DER).")


def get_name_attr(name: x509.Name, oid) -> str:
    try:
        attrs = name.get_attributes_for_oid(oid)
        if attrs:
            return attrs[0].value
    except Exception:
        pass
    return "-"


def subject_fields(name: x509.Name) -> dict:
    return {
        "Common Name":         get_name_attr(name, NameOID.COMMON_NAME),
        "emailAddress":        get_name_attr(name, NameOID.EMAIL_ADDRESS),
        "Organizational Unit": get_name_attr(name, NameOID.ORGANIZATIONAL_UNIT_NAME),
        "Organization":        get_name_attr(name, NameOID.ORGANIZATION_NAME),
        "Locality":            get_name_attr(name, NameOID.LOCALITY_NAME),
        "State or Province":   get_name_attr(name, NameOID.STATE_OR_PROVINCE_NAME),
        "Country":             get_name_attr(name, NameOID.COUNTRY_NAME),
    }


def issuer_fields(name: x509.Name) -> dict:
    return {
        "Issuer Common Name":      get_name_attr(name, NameOID.COMMON_NAME),
        "Issuer emailAddress":     get_name_attr(name, NameOID.EMAIL_ADDRESS),
        "Issuer Organization":     get_name_attr(name, NameOID.ORGANIZATION_NAME),
        "Issuer Locality":         get_name_attr(name, NameOID.LOCALITY_NAME),
        "Issuer State or Province": get_name_attr(name, NameOID.STATE_OR_PROVINCE_NAME),
        "Issuer Country":          get_name_attr(name, NameOID.COUNTRY_NAME),
    }


def format_name(name: x509.Name) -> str:
    parts = []
    for rdn in name.rdns:
        for attr in rdn:
            parts.append(f"{attr.oid._name}={attr.value}")
    return ", ".join(parts) if parts else "-"


def get_key_info(public_key):
    if isinstance(public_key, rsa.RSAPublicKey):
        return "RSA", str(public_key.key_size)
    if isinstance(public_key, dsa.DSAPublicKey):
        return "DSA", str(public_key.key_size)
    if isinstance(public_key, ec.EllipticCurvePublicKey):
        try:
            size = public_key.key_size
        except Exception:
            size = "-"
        return f"EC ({public_key.curve.name})", str(size)
    return public_key.__class__.__name__, "-"


def get_signature_algorithm(obj) -> str:
    try:
        sig_hash = obj.signature_hash_algorithm.name
    except Exception:
        sig_hash = "-"

    algo_name = "-"
    try:
        algo_name = obj.signature_algorithm_oid._name
    except Exception:
        pass

    if algo_name == "-":
        return sig_hash if sig_hash != "-" else "-"
    if sig_hash != "-":
        return f"{algo_name} ({sig_hash})"
    return algo_name


def compute_thumbprint(cert: x509.Certificate) -> str:
    try:
        fp = cert.fingerprint(hashes.SHA1())
        return fp.hex().upper()
    except Exception:
        return "-"


def get_validity_utc(obj):
    """
    Geeft (valid_from_iso, valid_to_iso) terug.
    Gebruikt *_utc als die bestaan (nieuwe cryptography),
    anders de oude not_valid_before / not_valid_after.
    """
    start = getattr(obj, "not_valid_before_utc", None)
    end = getattr(obj, "not_valid_after_utc", None)

    if start is None:
        start = obj.not_valid_before
    if end is None:
        end = obj.not_valid_after

    return start.isoformat(), end.isoformat()


def decode_file(path: Path) -> dict:
    """Decode een cert/CSR vanaf een bestandspad naar een gestructureerd dict."""
    data = path.read_bytes()
    return decode_file_from_bytes(data, path)


def decode_file_from_bytes(data: bytes, fake_path: Path) -> dict:
    """Decode een cert/CSR vanuit bytes (voor web uploads)."""
    obj_type, obj = load_cert_or_csr(data)

    subj_map = subject_fields(obj.subject)

    if obj_type == "cert":
        issuer_map = issuer_fields(obj.issuer)
        valid_from, valid_to = get_validity_utc(obj)
        serial = hex(obj.serial_number).upper().replace("X", "x")
        thumb = compute_thumbprint(obj)
        issuer_str = format_name(obj.issuer)
    else:
        issuer_map = None
        valid_from = "-"
        valid_to = "-"
        serial = "-"
        thumb = "-"
        issuer_str = "-"

    pub = obj.public_key()
    key_algo, key_size = get_key_info(pub)
    sig_algo = get_signature_algorithm(obj)

    props = {
        "Subject":        format_name(obj.subject),
        "Issuer":         issuer_str,
        "Valid From":     valid_from,
        "Valid To":       valid_to,
        "Key Size":       key_size,
        "Key Algorithm":  key_algo,
        "Sig. Algorithm": sig_algo,
        "Serial Number":  serial,
        "Thumbprint":     thumb,
    }

    info = {
        "filename": str(fake_path),
        "type": "Certificate" if obj_type == "cert" else "CSR",
        "subject": subj_map,
        "issuer": issuer_map,
        "properties": props,
    }
    return info


# ============================================================
#   EXPORT: BESTANDSGEBASEERDE FORMATS
# ============================================================

def export_json(info: dict, dest: Path):
    with dest.open("w", encoding="utf-8") as f:
        json.dump(info, f, indent=2, ensure_ascii=False)


def export_csv(info: dict, dest: Path):
    with dest.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Section", "Field", "Value"])
        for section_key, section_name in [
            ("subject", "Subject"),
            ("issuer", "Issuer"),
            ("properties", "Properties"),
        ]:
            section = info.get(section_key)
            if section is None:
                continue
            for field, value in section.items():
                writer.writerow([section_name, field, value])


def export_html(info: dict, dest: Path):
    def dict_to_html_table(title, mapping):
        if mapping is None:
            return f"<h2>{title}</h2><p>CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>"
        rows = "".join(
            f"<tr><th>{k}</th><td>{v}</td></tr>"
            for k, v in mapping.items()
        )
        return f"<h2>{title}</h2><table>{rows}</table>"

    subject_html = dict_to_html_table("Certificate Subject", info["subject"])
    issuer_html = dict_to_html_table("Certificate Issuer", info["issuer"])
    props_html = dict_to_html_table("Certificate Properties", info["properties"])

    html = f"""<!doctype html>
<html lang="nl">
<head>
<meta charset="utf-8">
<title>CyNiT Certificate Decoder Export</title>
<style>
body {{
  background: {BG_COLOR};
  color: {MAIN_COLOR};
  font-family: Arial, sans-serif;
}}
h1 {{
  color: {TITLE_COLOR};
}}
h2 {{
  color: {TITLE_COLOR};
  margin-top: 20px;
}}
table {{
  border-collapse: collapse;
  margin-bottom: 20px;
  min-width: 500px;
}}
th, td {{
  border: 1px solid #555;
  padding: 4px 8px;
}}
th {{
  background: {TABLE_COL1_BG};
  color: {TABLE_COL1_FG};
}}
td {{
  color: {TABLE_COL2_FG};
  background: {TABLE_COL2_BG};
}}
</style>
</head>
<body>
<h1>CyNiT Certificate Decoder Export</h1>
<p>Bestand: {info.get("filename", "")}</p>
<p>Type: {info.get("type", "")}</p>
{subject_html}
{issuer_html}
{props_html}
</body>
</html>
"""
    dest.write_text(html, encoding="utf-8")


def export_markdown(info: dict, dest: Path):
    def dict_to_md_table(title, mapping, issuer=False):
        if issuer and mapping is None:
            return f"## {title}\n\nCSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.\n\n"
        lines = [f"## {title}", "", "| Field | Value |", "| --- | --- |"]
        for k, v in mapping.items():
            lines.append(f"| {k} | {v} |")
        lines.append("")
        return "\n".join(lines) + "\n"

    md = []
    md.append(f"# CyNiT Certificate Decoder Export\n")
    md.append(f"**Bestand**: `{info.get('filename', '')}`  ")
    md.append(f"**Type**: {info.get('type', '')}\n")

    md.append(dict_to_md_table("Certificate Subject", info["subject"]))
    md.append(dict_to_md_table("Certificate Issuer", info["issuer"], issuer=True))
    md.append(dict_to_md_table("Certificate Properties", info["properties"]))

    dest.write_text("\n".join(md), encoding="utf-8")


def export_xlsx(info: dict, dest: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate"

    title_font = Font(color=TITLE_COLOR_XLSX, bold=True)
    yellow_fill = PatternFill(
        start_color=TABLE_COL1_BG.lstrip("#"),
        end_color=TABLE_COL1_BG.lstrip("#"),
        fill_type="solid"
    )
    main_font = Font(color=MAIN_COLOR_XLSX)

    row = 1
    ws["A1"] = "CyNiT Certificate Decoder Export"
    ws["A1"].font = title_font
    row += 2

    ws[f"A{row}"] = "Bestand"
    ws[f"B{row}"] = info.get("filename", "")
    row += 1
    ws[f"A{row}"] = "Type"
    ws[f"B{row}"] = info.get("type", "")
    row += 2

    def write_section(title, mapping):
        nonlocal row
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = title_font
        row += 1
        if mapping is None:
            ws[f"A{row}"] = "CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat."
            row += 2
            return
        ws[f"A{row}"] = "Field"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].fill = yellow_fill
        ws[f"B{row}"].fill = yellow_fill
        row += 1
        for k, v in mapping.items():
            ws[f"A{row}"] = k
            ws[f"A{row}"].fill = yellow_fill
            ws[f"B{row}"] = v
            ws[f"B{row}"].font = main_font
            row += 1
        row += 1

    write_section("Certificate Subject", info["subject"])
    write_section("Certificate Issuer", info["issuer"])
    write_section("Certificate Properties", info["properties"])

    wb.save(dest)


def export_all_formats(info: dict, base_path: Path):
    export_json(info, base_path.with_suffix(".json"))
    export_csv(info, base_path.with_suffix(".csv"))
    export_html(info, base_path.with_suffix(".html"))
    export_markdown(info, base_path.with_suffix(".md"))
    export_xlsx(info, base_path.with_suffix(".xlsx"))


# ============================================================
#   PYINSTALLER BUILD HELPER
# ============================================================

def run_pyinstaller_build():
    """
    Bouwt een onefile EXE met PyInstaller.
    In een al-gepackte EXE doen we dit niet (heeft geen zin).
    """
    if IS_FROZEN:
        return False, "Build EXE is niet beschikbaar in de reeds gebuilde EXE-versie."

    script_path = Path(__file__).resolve()
    try:
        cmd = [
            sys.executable,
            "-m",
            "PyInstaller",
            "--onefile",
            "--noconsole",
        ]

        # Icon op basis van master-logo -> tijdelijke ico op schijf
        ico_bytes = generate_ico_bytes()
        tmp_ico_path = None
        if ico_bytes is not None:
            try:
                tmp_ico_path = BASE_DIR / "images" / "_tmp_cynit_icon.ico"
                tmp_ico_path.parent.mkdir(exist_ok=True)
                tmp_ico_path.write_bytes(ico_bytes)
                cmd.append(f"--icon={tmp_ico_path.name}")
            except Exception:
                tmp_ico_path = None

        cmd += [
            "-y",
            script_path.name,
        ]

        subprocess.run(
            cmd,
            cwd=script_path.parent,
            check=True,
        )

        # tijdelijke ico opruimen
        if tmp_ico_path is not None and tmp_ico_path.exists():
            try:
                tmp_ico_path.unlink()
            except Exception:
                pass

        exe_name = script_path.stem + (".exe" if sys.platform.startswith("win") else "")
        dist_path = script_path.parent / "dist" / exe_name
        return True, f"PyInstaller build voltooid.\nEXE staat (waarschijnlijk) in:\n{dist_path}"
    except FileNotFoundError:
        return False, "PyInstaller lijkt niet geïnstalleerd.\nInstalleer met:\npip install pyinstaller"
    except subprocess.CalledProcessError as e:
        return False, f"PyInstaller is gefaald:\n{e}"


# ============================================================
#   TKINTER GUI (CyNiT LOOK & FEEL + ABOUT + PROFIEL DROPDOWN + TABEL)
# ============================================================

class CynitCertGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CyNiT Certificate / CSR Decoder")
        self.geometry("1100x780")

        self.bg_color = BG_COLOR
        self.fg_color = MAIN_COLOR
        self.title_color = TITLE_COLOR
        self.key_color = TABLE_COL1_BG
        self.value_color = TABLE_COL2_FG

        self.base_font = (UI_CFG.get("font_main", "Consolas"), 12)
        self.button_font = (UI_CFG.get("font_buttons", "Segoe UI"), 11, "bold")
        self.label_font = (UI_CFG.get("font_main", "Consolas"), 11)

        self.configure(bg=self.bg_color)

        self.current_info = None
        self.current_path = None

        self.export_buttons = []
        self.logo_img = None
        self.btn_build_exe = None  # referentie naar build-knop

        self.profile_var = tk.StringVar(value=ACTIVE_PROFILE_NAME)

        self._build_menu()
        self._build_header()   # logo links + profiel rechts
        self._build_gui()

    # ---- Menu (alleen Help nu) ----
    def _build_menu(self):
        menubar = tk.Menu(self)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About CyNiT Cert Tool", command=self.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.config(menu=menubar)

    def _build_header(self):
        """Bovenste balk met links logo en rechts profielkeuze."""
        header = tk.Frame(self, bg=self.bg_color)
        header.pack(fill=tk.X, padx=10, pady=(10, 0))

        # LINKS: logo
        left = tk.Frame(header, bg=self.bg_color)
        left.pack(side=tk.LEFT, anchor="w")

        logo_path = LOGO_BASE
        if logo_path.exists():
            try:
                img = Image.open(logo_path)
                max_h = int(UI_CFG.get("logo_max_height", 80))
                if img.height > 0:
                    scale = max_h / img.height
                else:
                    scale = 1.0
                new_w = int(img.width * scale)
                new_h = int(img.height * scale)
                img_resized = img.resize((new_w, new_h), Image.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(img_resized)
                logo_label = tk.Label(left, image=self.logo_img, bg=self.bg_color)
                logo_label.pack(side=tk.LEFT)
            except Exception:
                pass

        # RECHTS: profielbar
        right = tk.Frame(header, bg=self.bg_color)
        right.pack(side=tk.RIGHT, anchor="e")

        lbl = tk.Label(
            right,
            text="Profiel:",
            bg=self.bg_color,
            fg=self.fg_color,
            font=self.label_font
        )
        lbl.pack(side=tk.LEFT)

        profile_names = list(CONFIG["profiles"].keys())
        self.profile_var.set(CONFIG["active_profile"])

        option = tk.OptionMenu(right, self.profile_var, *profile_names, command=self.on_profile_selected)
        option.configure(
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            activebackground=BUTTON_BG,
            activeforeground=BUTTON_FG,
            relief=tk.RAISED,
            bd=3
        )
        option["menu"].configure(
            bg=self.bg_color,
            fg=self.fg_color
        )
        option.pack(side=tk.LEFT, padx=5)

        btn_new = tk.Button(
            right,
            text="Nieuw profiel…",
            command=self.create_profile_dialog,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            activebackground=BUTTON_BG,
            activeforeground=BUTTON_FG,
            font=self.button_font,
            relief=tk.RAISED,
            bd=3
        )
        btn_new.pack(side=tk.LEFT, padx=5)

        lbl_active = tk.Label(
            right,
            text=f"Actief: {CONFIG['active_profile']}",
            bg=self.bg_color,
            fg=self.fg_color,
            font=("Consolas", 9)
        )
        lbl_active.pack(side=tk.LEFT, padx=10)

    def on_profile_selected(self, name):
        """Als de user een ander profiel kiest in de dropdown."""
        if not name or name == CONFIG["active_profile"]:
            return
        CONFIG["active_profile"] = name
        save_config(CONFIG)
        restart_program()

    def create_profile_dialog(self):
        """Dialoog om een nieuw profiel aan te maken op basis van huidig profiel."""
        base_profile = CONFIG["profiles"][CONFIG["active_profile"]]

        dlg = tk.Toplevel(self)
        dlg.title("Nieuw profiel")
        dlg.configure(bg=self.bg_color)
        dlg.resizable(False, False)

        tk.Label(
            dlg,
            text="Profielnaam:",
            bg=self.bg_color,
            fg=self.fg_color,
            font=self.label_font
        ).grid(row=0, column=0, sticky="w", padx=10, pady=5)

        name_var = tk.StringVar()
        tk.Entry(
            dlg,
            textvariable=name_var,
            bg="#111111",
            fg=self.fg_color,
            insertbackground=self.fg_color,
            font=self.base_font,
            width=30
        ).grid(row=0, column=1, sticky="w", padx=10, pady=5)

        row = 2
        entry_vars = {}

        def add_field(label, key_path, default_value):
            nonlocal row
            tk.Label(
                dlg,
                text=label,
                bg=self.bg_color,
                fg=self.fg_color,
                font=self.label_font
            ).grid(row=row, column=0, sticky="w", padx=10, pady=2)

            var = tk.StringVar(value=str(default_value))
            tk.Entry(
                dlg,
                textvariable=var,
                bg="#111111",
                fg=self.fg_color,
                insertbackground=self.fg_color,
                font=self.base_font,
                width=30
            ).grid(row=row, column=1, sticky="w", padx=10, pady=2)
            entry_vars[key_path] = var
            row += 1

        colors = base_profile["colors"]
        paths = base_profile["paths"]
        ui = base_profile["ui"]

        add_field("Achtergrond", "colors.background", colors["background"])
        add_field("Algemene tekstkleur", "colors.general_fg", colors["general_fg"])
        add_field("Titel kleur", "colors.title", colors["title"])
        add_field("Kolom 1 achtergrond", "colors.table_col1_bg", colors["table_col1_bg"])
        add_field("Kolom 1 voorgrond", "colors.table_col1_fg", colors["table_col1_fg"])
        add_field("Kolom 2 achtergrond", "colors.table_col2_bg", colors["table_col2_bg"])
        add_field("Kolom 2 voorgrond", "colors.table_col2_fg", colors["table_col2_fg"])
        add_field("Knoppen achtergrond", "colors.button_bg", colors["button_bg"])
        add_field("Knoppen voorgrond", "colors.button_fg", colors["button_fg"])

        add_field("Logo pad", "paths.logo", paths["logo"])
        add_field("Help pad (ABOUT)", "paths.help", paths["help"])

        add_field("Logo max hoogte (px)", "ui.logo_max_height", ui.get("logo_max_height", 80))
        add_field("Hoofdlettertype", "ui.font_main", ui.get("font_main", "Consolas"))
        add_field("Knoppen lettertype", "ui.font_buttons", ui.get("font_buttons", "Segoe UI"))

        btn_frame = tk.Frame(dlg, bg=self.bg_color)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)

        def on_cancel():
            dlg.destroy()

        def on_save():
            new_name = name_var.get().strip()
            if not new_name:
                messagebox.showerror("Profiel", "Gelieve een profielnaam in te vullen.", parent=dlg)
                return
            if new_name in CONFIG["profiles"]:
                messagebox.showerror("Profiel", f"Profiel '{new_name}' bestaat al.", parent=dlg)
                return

            new_profile = copy.deepcopy(base_profile)

            def set_in_dict(d, path, value):
                parts = path.split(".")
                cur = d
                for p in parts[:-1]:
                    cur = cur.setdefault(p, {})
                cur[parts[-1]] = value

            for key_path, var in entry_vars.items():
                val = var.get()
                if key_path == "ui.logo_max_height":
                    try:
                        val = int(val)
                    except ValueError:
                        val = ui.get("logo_max_height", 80)
                set_in_dict(new_profile, key_path, val)

            CONFIG["profiles"][new_name] = new_profile
            CONFIG["active_profile"] = new_name
            save_config(CONFIG)

            messagebox.showinfo(
                "Profiel aangemaakt",
                f"Profiel '{new_name}' is aangemaakt en geactiveerd.\n"
                f"De applicatie wordt nu herstart met dit profiel.",
                parent=dlg
            )
            dlg.destroy()
            restart_program()

        tk.Button(
            btn_frame,
            text="Annuleren",
            command=on_cancel,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            activebackground=BUTTON_BG,
            activeforeground=BUTTON_FG,
            font=self.button_font,
            width=12,
            relief=tk.RAISED,
            bd=3
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="Opslaan",
            command=on_save,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            activebackground=BUTTON_BG,
            activeforeground=BUTTON_FG,
            font=self.button_font,
            width=12,
            relief=tk.RAISED,
            bd=3
        ).pack(side=tk.LEFT, padx=5)

        dlg.grab_set()
        dlg.transient(self)
        dlg.focus_set()

    def show_about(self):
        """Toont ABOUT.md in een nieuw venster."""
        try:
            text = ABOUT_MD.read_text(encoding="utf-8")
        except Exception:
            text = "ABOUT.md kon niet worden gelezen."

        win = tk.Toplevel(self)
        win.title("About CyNiT Certificate / CSR Decoder")
        win.geometry("700x500")
        win.configure(bg=self.bg_color)

        txt = scrolledtext.ScrolledText(
            win,
            wrap=tk.WORD,
            font=(UI_CFG.get("font_main", "Consolas"), 11),
            bg=self.bg_color,
            fg=self.fg_color,
            insertbackground=self.fg_color,
        )
        txt.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        txt.insert("1.0", text)
        txt.config(state=tk.DISABLED)

    def _build_gui(self):
        top_frame = tk.Frame(self, bg=self.bg_color)
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        left_frame = tk.Frame(top_frame, bg=self.bg_color)
        left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        btn_open = tk.Button(
            left_frame,
            text="Bestand kiezen…",
            command=self.choose_file,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            activebackground=BUTTON_BG,
            activeforeground=BUTTON_FG,
            font=self.button_font,
            relief=tk.RAISED,
            bd=3
        )
        btn_open.pack(side=tk.LEFT)

        self.lbl_file = tk.Label(
            left_frame,
            text="Geen bestand geselecteerd",
            bg=self.bg_color,
            fg=self.fg_color,
            anchor="w",
            font=self.label_font
        )
        self.lbl_file.pack(side=tk.LEFT, padx=10)

        right_frame = tk.Frame(top_frame, bg=self.bg_color)
        right_frame.pack(side=tk.RIGHT, anchor="ne")

        btn_webui = tk.Button(
            right_frame,
            text="Open Web UI",
            command=self.open_web_ui,
            bg=BUTTON_BG,
            fg=BUTTON_FG,
            activebackground=BUTTON_BG,
            activeforeground=BUTTON_FG,
            font=self.button_font,
            width=20,
            relief=tk.RAISED,
            bd=3
        )
        btn_webui.pack(side=tk.TOP, pady=2, anchor="e")

        if not IS_FROZEN:
            self.btn_build_exe = tk.Button(
                right_frame,
                text="Build EXE (PyInstaller)",
                command=self.build_exe,
                bg=BUTTON_BG,
                fg=BUTTON_FG,
                activebackground=BUTTON_BG,
                activeforeground=BUTTON_FG,
                font=self.button_font,
                width=20,
                relief=tk.RAISED,
                bd=3
            )
            self.btn_build_exe.pack(side=tk.TOP, pady=2, anchor="e")

        def make_export_button(text, fmt):
            btn = tk.Button(
                right_frame,
                text=text,
                command=lambda f=fmt: self.export_current(f),
                bg=BUTTON_BG,
                fg=BUTTON_FG,
                activebackground=BUTTON_BG,
                activeforeground=BUTTON_FG,
                font=self.button_font,
                state=tk.DISABLED,
                width=20,
                relief=tk.RAISED,
                bd=3
            )
            btn.pack(side=tk.TOP, pady=2, anchor="e")
            self.export_buttons.append(btn)
            return btn

        make_export_button("Export JSON", "json")
        make_export_button("Export CSV", "csv")
        make_export_button("Export XLSX", "xlsx")
        make_export_button("Export HTML", "html")
        make_export_button("Export Markdown", "md")
        make_export_button("Export ALL", "all")

        # TABEL-achtige view met scroll
        table_frame = tk.Frame(self, bg=self.bg_color)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.table_canvas = tk.Canvas(
            table_frame,
            bg=self.bg_color,
            highlightthickness=0
        )
        self.table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.table_canvas.yview
        )
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.table_canvas.configure(yscrollcommand=scrollbar.set)

        self.table_inner = tk.Frame(self.table_canvas, bg=self.bg_color)
        self.table_canvas.create_window((0, 0), window=self.table_inner, anchor="nw")

        def on_configure(event):
            self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all"))

        self.table_inner.bind("<Configure>", on_configure)

    def _set_export_buttons_state(self, state: str):
        for btn in self.export_buttons:
            btn.config(state=state)

    def open_web_ui(self):
        start_web_in_background()
        url = f"http://{WEB_HOST}:{WEB_PORT}/"
        try:
            webbrowser.open(url)
        except Exception:
            messagebox.showinfo("Web UI", f"Open deze URL in je browser:\n{url}")

    def build_exe(self):
        ok, msg = run_pyinstaller_build()
        if ok:
            messagebox.showinfo("PyInstaller", msg)
            if self.btn_build_exe is not None:
                self.btn_build_exe.pack_forget()
                self.btn_build_exe = None
        else:
            messagebox.showerror("PyInstaller", msg)

    def choose_file(self):
        filetypes = [
            ("Alle ondersteunde bestanden", "*.crt *.cer *.pem *.csr"),
            ("Certificates", "*.crt *.cer *.pem"),
            ("Certificate Signing Requests", "*.csr"),
            ("Alle bestanden", "*.*"),
        ]
        filename = filedialog.asksaveasfilename(
            title="Kies een certificaat of CSR",
            filetypes=filetypes
        )
        if not filename:
            return

        path = Path(filename)
        self.lbl_file.config(text=str(path))

        try:
            info = decode_file(path)
        except Exception as e:
            messagebox.showerror("Fout", f"Kon bestand niet decoderen:\n{e}")
            self._set_export_buttons_state(tk.DISABLED)
            return

        self.current_info = info
        self.current_path = path
        self._set_export_buttons_state(tk.NORMAL)
        self.show_info(info)

    def _clear_table(self):
        for w in self.table_inner.winfo_children():
            w.destroy()

    def show_info(self, info: dict):
        self._clear_table()

        row = 0

        def add_section_title(title):
            nonlocal row
            lbl = tk.Label(
                self.table_inner,
                text=title,
                bg=self.bg_color,
                fg=self.title_color,
                font=(UI_CFG.get("font_main", "Consolas"), 13, "bold")
            )
            lbl.grid(row=row, column=0, columnspan=2, sticky="w", pady=(10, 2))
            row += 1

        def add_separator():
            nonlocal row
            sep = tk.Label(
                self.table_inner,
                text="─" * 80,
                bg=self.bg_color,
                fg=self.fg_color,
                font=(UI_CFG.get("font_main", "Consolas"), 9)
            )
            sep.grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 5))
            row += 1

        def add_kv(key, value):
            nonlocal row
            key_lbl = tk.Label(
                self.table_inner,
                text=key,
                bg=TABLE_COL1_BG,
                fg=TABLE_COL1_FG,
                font=self.label_font,
                anchor="w"
            )
            key_lbl.grid(row=row, column=0, sticky="nsew", padx=(0, 2), pady=1, ipadx=4, ipady=2)

            val_lbl = tk.Label(
                self.table_inner,
                text=value,
                bg=TABLE_COL2_BG,
                fg=TABLE_COL2_FG,
                font=self.label_font,
                anchor="w",
                justify="left",
                wraplength=800
            )
            val_lbl.grid(row=row, column=1, sticky="nsew", padx=(2, 0), pady=1, ipadx=4, ipady=2)

            self.table_inner.grid_columnconfigure(0, weight=1)
            self.table_inner.grid_columnconfigure(1, weight=2)
            row += 1

        # Bestand / Type
        add_section_title(f"Bestand: {info.get('filename', '')}")
        add_kv("Type", info.get("type", ""))
        add_separator()

        # Subject
        add_section_title("Certificate Subject")
        for k, v in info["subject"].items():
            add_kv(k, v)
        add_separator()

        # Issuer
        add_section_title("Certificate Issuer")
        if info["issuer"] is None:
            add_kv("Issuer", "CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.")
        else:
            for k, v in info["issuer"].items():
                add_kv(k, v)
        add_separator()

        # Properties
        add_section_title("Certificate Properties")
        for k, v in info["properties"].items():
            add_kv(k, v)

    def export_current(self, fmt: str):
        if not self.current_info:
            messagebox.showwarning("Geen data", "Er is nog geen certificaat/CSR geladen.")
            return

        if fmt == "all":
            base = filedialog.asksaveasfilename(
                title="Kies basenaam voor ALL export (zonder extensie)",
                defaultextension=".json",
                filetypes=[("JSON", "*.json"), ("Alle bestanden", "*.*")]
            )
            if not base:
                return
            base_path = Path(base).with_suffix("")
            try:
                export_all_formats(self.current_info, base_path)
            except Exception as e:
                messagebox.showerror("Export-fout", f"Export is mislukt:\n{e}")
                return
            messagebox.showinfo(
                "Export voltooid",
                "Alle formaten zijn aangemaakt:\n\n"
                f"{base_path.with_suffix('.json')}\n"
                f"{base_path.with_suffix('.csv')}\n"
                f"{base_path.with_suffix('.xlsx')}\n"
                f"{base_path.with_suffix('.html')}\n"
                f"{base_path.with_suffix('.md')}\n"
            )
            return

        ext_map = {
            "json": (".json", [("JSON", "*.json"), ("Alle bestanden", "*.*")]),
            "csv":  (".csv", [("CSV", "*.csv"), ("Alle bestanden", "*.*")]),
            "xlsx": (".xlsx", [("Excel XLSX", "*.xlsx"), ("Alle bestanden", "*.*")]),
            "html": (".html", [("HTML", "*.html"), ("Alle bestanden", "*.*")]),
            "md":   (".md", [("Markdown", "*.md"), ("Alle bestanden", "*.*")]),
        }

        default_ext, filetypes = ext_map[fmt]

        filename = filedialog.asksaveasfilename(
            title=f"Export {fmt.upper()} opslaan als",
            defaultextension=default_ext,
            filetypes=filetypes
        )
        if not filename:
            return

        dest = Path(filename)

        try:
            if fmt == "json":
                export_json(self.current_info, dest)
            elif fmt == "csv":
                export_csv(self.current_info, dest)
            elif fmt == "xlsx":
                export_xlsx(self.current_info, dest)
            elif fmt == "html":
                export_html(self.current_info, dest)
            elif fmt == "md":
                export_markdown(self.current_info, dest)
        except Exception as e:
            messagebox.showerror("Export-fout", f"Export is mislukt:\n{e}")
            return

        messagebox.showinfo("Export voltooid", f"Export opgeslagen als:\n{dest}")


# ============================================================
#   FLASK WEB UI (CyNiT LOOK & FEEL + HAMBURGER + WAFFLE EXPORT + PROFIELBEHEER)
# ============================================================

app = Flask(__name__)

WEB_TEMPLATE = f"""
<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <title>CyNiT Certificate / CSR Decoder</title>
  <link rel="icon" type="image/x-icon" href="/favicon.ico">
  <style>
    body {{
      background: {BG_COLOR};
      color: {MAIN_COLOR};
      font-family: Arial, sans-serif;
      margin: 0;
    }}
    .topbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 8px 16px;
      background: #111111;
      border-bottom: 1px solid #333333;
    }}
    .topbar-left {{
      display: flex;
      align-items: center;
      gap: 10px;
    }}
    .topbar-right {{
      display: flex;
      align-items: center;
      gap: 10px;
    }}
    .icon-button {{
      background: {BUTTON_BG};
      color: {MAIN_COLOR};
      border: 1px solid {MAIN_COLOR};
      border-radius: 4px;
      padding: 4px 8px;
      font-size: 18px;
      cursor: pointer;
      box-shadow: 0 2px 4px rgba(0,0,0,0.6);
    }}
    .icon-button:hover {{
      background: #222222;
    }}
    .logo {{
      max-height: {UI_CFG.get("logo_max_height", 80)}px;
    }}
    .page {{
      padding: 20px;
    }}
    h1, h2 {{
      color: {TITLE_COLOR};
    }}
    table {{
      border-collapse: collapse;
      margin-bottom: 20px;
      min-width: 500px;
    }}
    th, td {{
      border: 1px solid #555;
      padding: 4px 8px;
    }}
    th {{
      background: {TABLE_COL1_BG};
      color: {TABLE_COL1_FG};
    }}
    td {{
      color: {TABLE_COL2_FG};
      background: {TABLE_COL2_BG};
    }}
    .error {{
      color: #ff0000;
      font-weight: bold;
    }}
    .section-title {{
      margin-top: 20px;
    }}
    label, button, a {{
      color: {MAIN_COLOR};
    }}
    input[type="file"] {{
      color: {MAIN_COLOR};
    }}
    button {{
      background: {BUTTON_BG};
      border: 1px solid {MAIN_COLOR};
      padding: 5px 10px;
      cursor: pointer;
      margin-right: 5px;
      box-shadow: 0 2px 4px rgba(0,0,0,0.6);
    }}
    button:hover {{
      background: #222222;
    }}
    .menu-panel {{
      position: absolute;
      top: 50px;
      left: 10px;
      background: #111111;
      border: 1px solid #333333;
      padding: 10px;
      display: none;
      z-index: 10;
      min-width: 220px;
    }}
    .menu-panel a, .menu-panel button {{
      display: block;
      width: 100%;
      text-align: left;
      margin: 3px 0;
    }}
    .export-panel {{
      position: absolute;
      top: 50px;
      right: 10px;
      background: #111111;
      border: 1px solid #333333;
      padding: 10px;
      display: none;
      z-index: 10;
      min-width: 220px;
    }}
  </style>
</head>
<body>
  <div class="topbar">
    <div class="topbar-left">
      <button class="icon-button" onclick="toggleNavMenu()">☰</button>
      <img src="/logo.png" alt="CyNiT Logo" class="logo">
      <span>CyNiT Certificate / CSR Decoder</span>
    </div>
    <div class="topbar-right">
      {{% if info %}}
      <button class="icon-button" onclick="toggleExportMenu()" title="Exports">▦</button>
      {{% endif %}}
    </div>
  </div>

  <div id="navMenu" class="menu-panel">
    <a href="/">🏠 Home</a>
    <a href="/about" target="_blank">ℹ️ About</a>
    <a href="/profiles" target="_blank">🎨 Profielen beheren</a>
    <form method="post" action="/start-gui">
      <button type="submit">🖥️ Start GUI</button>
    </form>
    <form method="post" action="/build-exe">
      <button type="submit">📦 Build EXE (PyInstaller)</button>
    </form>
  </div>

  {{% if info %}}
  <div id="exportMenu" class="export-panel">
    <form method="get" action="/download/json">
      <button type="submit">⬇ JSON</button>
    </form>
    <form method="get" action="/download/csv">
      <button type="submit">⬇ CSV</button>
    </form>
    <form method="get" action="/download/xlsx">
      <button type="submit">⬇ XLSX</button>
    </form>
    <form method="get" action="/download/html">
      <button type="submit">⬇ HTML</button>
    </form>
    <form method="get" action="/download/md">
      <button type="submit">⬇ Markdown</button>
    </form>
  </div>
  {{% endif %}}

  <div class="page">
    <form method="post" enctype="multipart/form-data">
      <label>Upload certificaat of CSR:
        <input type="file" name="file">
      </label>
      <button type="submit">Decode</button>
    </form>

    {{% if gui_msg %}}
      <p>{{{{ gui_msg }}}}</p>
    {{% endif %}}
    {{% if build_msg %}}
      <p>{{{{ build_msg }}}}</p>
    {{% endif %}}

    {{% if error %}}
      <p class="error">{{{{ error }}}}</p>
    {{% endif %}}

    {{% if info %}}
      <h2>Resultaat</h2>
      <p><strong>Bestand:</strong> {{{{ info.filename }}}}</p>
      <p><strong>Type:</strong> {{{{ info.type }}}}</p>

      <h3 class="section-title">Certificate Subject</h3>
      <table>
        <tbody>
          {{% for k, v in info.subject.items() %}}
          <tr>
            <th>{{{{ k }}}}</th><td>{{{{ v }}}}</td>
          </tr>
          {{% endfor %}}
        </tbody>
      </table>

      <h3 class="section-title">Certificate Issuer</h3>
      {{% if info.issuer %}}
        <table>
          <tbody>
            {{% for k, v in info.issuer.items() %}}
            <tr>
              <th>{{{{ k }}}}</th><td>{{{{ v }}}}</td>
            </tr>
            {{% endfor %}}
          </tbody>
        </table>
      {{% else %}}
        <p>CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>
      {{% endif %}}

      <h3 class="section-title">Certificate Properties</h3>
      <table>
        <tbody>
          {{% for k, v in info.properties.items() %}}
          <tr>
            <th>{{{{ k }}}}</th><td>{{{{ v }}}}</td>
          </tr>
          {{% endfor %}}
        </tbody>
      </table>

    {{% endif %}}
  </div>

  <script>
    function toggleNavMenu() {{
      var m = document.getElementById("navMenu");
      if (m.style.display === "block") {{
        m.style.display = "none";
      }} else {{
        m.style.display = "block";
      }}
    }}
    function toggleExportMenu() {{
      var m = document.getElementById("exportMenu");
      if (!m) return;
      if (m.style.display === "block") {{
        m.style.display = "none";
      }} else {{
        m.style.display = "block";
      }}
    }}
    // Sluit menu's bij klik buiten
    document.addEventListener("click", function(e) {{
      var nav = document.getElementById("navMenu");
      var exp = document.getElementById("exportMenu");
      var target = e.target;
      if (nav && !nav.contains(target) && !target.closest(".topbar-left")) {{
        nav.style.display = "none";
      }}
      if (exp && !exp.contains(target) && !target.closest(".topbar-right")) {{
        exp.style.display = "none";
      }}
    }});
  </script>
</body>
</html>
"""

PROFILES_TEMPLATE = f"""
<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <title>CyNiT Profielbeheer</title>
  <link rel="icon" type="image/x-icon" href="/favicon.ico">
  <style>
    body {{
      background: {BG_COLOR};
      color: {MAIN_COLOR};
      font-family: Arial, sans-serif;
      margin: 20px;
    }}
    h1, h2, h3 {{
      color: {TITLE_COLOR};
    }}
    table {{
      border-collapse: collapse;
      margin-bottom: 20px;
    }}
    th, td {{
      border: 1px solid #555;
      padding: 4px 8px;
    }}
    label, button, a, input {{
      color: {MAIN_COLOR};
    }}
    input[type="text"] {{
      background: #111111;
      border: 1px solid #555;
      padding: 3px 5px;
    }}
    button {{
      background: {BUTTON_BG};
      border: 1px solid {MAIN_COLOR};
      padding: 5px 10px;
      cursor: pointer;
      margin: 2px;
      box-shadow: 0 2px 4px rgba(0,0,0,0.6);
    }}
    button:hover {{
      background: #222222;
    }}
    .small {{
      font-size: 0.9em;
      color: {MAIN_COLOR};
    }}
  </style>
</head>
<body>
  <h1>CyNiT Profielbeheer</h1>
  <p>Actief profiel: <strong>{{{{ active_profile }}}}</strong></p>
  <p><a href="/">Terug naar hoofdtool</a></p>

  <h2>Bestaande profielen</h2>
  <ul>
    {{% for name, prof in profiles.items() %}}
      <li>
        <strong>{{{{ name }}}}</strong>
        {{% if name == active_profile %}}
          (actief)
        {{% else %}}
          <form method="post" action="/profiles/activate" style="display:inline;">
            <input type="hidden" name="profile_name" value="{{{{ name }}}}">
            <button type="submit">Activeer</button>
          </form>
        {{% endif %}}
      </li>
    {{% endfor %}}
  </ul>

  <h2>Nieuw profiel aanmaken</h2>
  <p class="small">Nieuwe profielen starten vanuit het huidige actieve profiel. Pas hieronder de waarden aan en klik op Opslaan.</p>
  <form method="post" action="/profiles/create">
    <table>
      <tr>
        <th>Veld</th>
        <th>Waarde</th>
      </tr>
      <tr>
        <td>Profielnaam</td>
        <td><input type="text" name="name" value="" size="30" required></td>
      </tr>

      <tr><th colspan="2">Kleuren</th></tr>
      <tr>
        <td>Achtergrond (background)</td>
        <td><input type="text" name="colors.background" value="{{{{ base_profile.colors.background }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Algemene tekstkleur (general_fg)</td>
        <td><input type="text" name="colors.general_fg" value="{{{{ base_profile.colors.general_fg }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Titel kleur (title)</td>
        <td><input type="text" name="colors.title" value="{{{{ base_profile.colors.title }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Kolom 1 achtergrond (table_col1_bg)</td>
        <td><input type="text" name="colors.table_col1_bg" value="{{{{ base_profile.colors.table_col1_bg }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Kolom 1 voorgrond (table_col1_fg)</td>
        <td><input type="text" name="colors.table_col1_fg" value="{{{{ base_profile.colors.table_col1_fg }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Kolom 2 achtergrond (table_col2_bg)</td>
        <td><input type="text" name="colors.table_col2_bg" value="{{{{ base_profile.colors.table_col2_bg }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Kolom 2 voorgrond (table_col2_fg)</td>
        <td><input type="text" name="colors.table_col2_fg" value="{{{{ base_profile.colors.table_col2_fg }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Knoppen achtergrond (button_bg)</td>
        <td><input type="text" name="colors.button_bg" value="{{{{ base_profile.colors.button_bg }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Knoppen voorgrond (button_fg)</td>
        <td><input type="text" name="colors.button_fg" value="{{{{ base_profile.colors.button_fg }}}}" size="30"></td>
      </tr>

      <tr><th colspan="2">Paden</th></tr>
      <tr>
        <td>Logo pad</td>
        <td><input type="text" name="paths.logo" value="{{{{ base_profile.paths.logo }}}}" size="50"></td>
      </tr>
      <tr>
        <td>Help pad (ABOUT)</td>
        <td><input type="text" name="paths.help" value="{{{{ base_profile.paths.help }}}}" size="50"></td>
      </tr>

      <tr><th colspan="2">UI</th></tr>
      <tr>
        <td>Logo max hoogte (px)</td>
        <td><input type="text" name="ui.logo_max_height" value="{{{{ base_profile.ui.logo_max_height }}}}" size="10"></td>
      </tr>
      <tr>
        <td>Hoofdlettertype (font_main)</td>
        <td><input type="text" name="ui.font_main" value="{{{{ base_profile.ui.font_main }}}}" size="30"></td>
      </tr>
      <tr>
        <td>Knoppen lettertype (font_buttons)</td>
        <td><input type="text" name="ui.font_buttons" value="{{{{ base_profile.ui.font_buttons }}}}" size="30"></td>
      </tr>
    </table>

    <button type="submit">Nieuw profiel opslaan &amp; activeren</button>
  </form>
</body>
</html>
"""


def info_to_inmemory_exports(info: dict):
    json_str = json.dumps(info, indent=2, ensure_ascii=False)

    csv_buf = []
    csv_buf.append("Section;Field;Value")
    for section_key, section_name in [
        ("subject", "Subject"),
        ("issuer", "Issuer"),
        ("properties", "Properties"),
    ]:
        section = info.get(section_key)
        if section is None:
            continue
        for field, value in section.items():
            value_str = str(value).replace(";", ",")
            csv_buf.append(f"{section_name};{field};{value_str}")
    csv_str = "\n".join(csv_buf)

    def dict_to_html_table(title, mapping, issuer=False):
        if issuer and mapping is None:
            return f"<h3>{title}</h3><p>CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>"
        rows = "".join(
            f"<tr><th>{k}</th><td>{v}</td></tr>"
            for k, v in mapping.items()
        )
        return f"<h3>{title}</h3><table>{rows}</table>"

    html_frag = []
    html_frag.append(dict_to_html_table("Certificate Subject", info["subject"]))
    html_frag.append(dict_to_html_table("Certificate Issuer", info["issuer"], issuer=True))
    html_frag.append(dict_to_html_table("Certificate Properties", info["properties"]))
    html_str = "\n".join(html_frag)

    def dict_to_md_table(title, mapping, issuer=False):
        if issuer and mapping is None:
            return f"## {title}\n\nCSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.\n"
        lines = [f"## {title}", "", "| Field | Value |", "| --- | --- |"]
        for k, v in mapping.items():
            lines.append(f"| {k} | {v} |")
        lines.append("")
        return "\n".join(lines)

    md_sections = [
        f"# CyNiT Certificate Decoder Export",
        f"**Bestand**: `{info.get('filename', '')}`  ",
        f"**Type**: {info.get('type', '')}",
        "",
        dict_to_md_table("Certificate Subject", info["subject"]),
        dict_to_md_table("Certificate Issuer", info["issuer"], issuer=True),
        dict_to_md_table("Certificate Properties", info["properties"]),
    ]
    md_str = "\n".join(md_sections)

    return {
        "json": json_str,
        "csv": csv_str,
        "html": html_str,
        "md": md_str,
    }


def start_gui_subprocess():
    """
    Start de GUI in een apart proces.

    - In 'gewone' .py modus: python cert_tool.py --gui-only
    - In PyInstaller EXE modus: cert_tool.exe --gui-only
    """
    try:
        if IS_FROZEN:
            # We draaien als EXE -> start dezelfde EXE met --gui-only
            exe_path = Path(sys.executable)
            subprocess.Popen(
                [str(exe_path), "--gui-only"],
                cwd=exe_path.parent,
            )
        else:
            # We draaien als .py script
            script_path = Path(__file__).resolve()
            subprocess.Popen(
                [sys.executable, str(script_path), "--gui-only"],
                cwd=script_path.parent,
            )
        return True, "GUI werd opgestart (indien toegestaan door het systeem)."
    except Exception as e:
        return False, f"Kon GUI niet starten: {e}"


# ------------------- WEB ROUTES -----------------------

@app.route("/", methods=["GET", "POST"])
def web_index():
    global LAST_INFO
    error = None
    info_obj = None

    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            error = "Geen bestand geselecteerd."
        else:
            try:
                data = file.read()
                fake_path = Path(file.filename)
                info = decode_file_from_bytes(data, fake_path)
                info_obj = info
                LAST_INFO = info
            except Exception as e:
                error = f"Fout bij decoderen: {e}"

    return render_template_string(
        WEB_TEMPLATE,
        error=error,
        info=info_obj,
        gui_msg=None,
        build_msg=None,
    )


@app.route("/start-gui", methods=["POST"])
def web_start_gui():
    ok, msg = start_gui_subprocess()
    return render_template_string(
        WEB_TEMPLATE,
        error=None,
        info=None,
        gui_msg=msg,
        build_msg=None,
    )


@app.route("/build-exe", methods=["POST"])
def web_build_exe():
    ok, msg = run_pyinstaller_build()
    return render_template_string(
        WEB_TEMPLATE,
        error=None,
        info=None,
        gui_msg=None,
        build_msg=msg,
    )


@app.route("/download/<fmt>", methods=["GET"])
def web_download(fmt):
    global LAST_INFO
    if LAST_INFO is None:
        return make_response("Nog geen certificaat/CSR gedecodeerd in deze sessie.", 400)

    base_name = Path(LAST_INFO.get("filename", "certificate")).stem or "certificate"

    if fmt == "json":
        content = json.dumps(LAST_INFO, indent=2, ensure_ascii=False)
        resp = make_response(content)
        resp.headers["Content-Type"] = "application/json; charset=utf-8"
        resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.json"'
        return resp

    if fmt == "csv":
        exports = info_to_inmemory_exports(LAST_INFO)
        content = exports["csv"]
        resp = make_response(content)
        resp.headers["Content-Type"] = "text/csv; charset=utf-8"
        resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.csv"'
        return resp

    if fmt == "html":
        tmp = BASE_DIR / "_tmp_export.html"
        export_html(LAST_INFO, tmp)
        content = tmp.read_text(encoding="utf-8")
        tmp.unlink(missing_ok=True)
        resp = make_response(content)
        resp.headers["Content-Type"] = "text/html; charset=utf-8"
        resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.html"'
        return resp

    if fmt == "md":
        tmp = BASE_DIR / "_tmp_export.md"
        export_markdown(LAST_INFO, tmp)
        content = tmp.read_text(encoding="utf-8")
        tmp.unlink(missing_ok=True)
        resp = make_response(content)
        resp.headers["Content-Type"] = "text/markdown; charset=utf-8"
        resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.md"'
        return resp

    if fmt == "xlsx":
        tmp = BASE_DIR / "_tmp_export.xlsx"
        export_xlsx(LAST_INFO, tmp)
        with tmp.open("rb") as f:
            data = f.read()
        tmp.unlink(missing_ok=True)
        buf = BytesIO(data)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{base_name}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return make_response("Onbekend exporttype.", 400)


@app.route("/favicon.ico")
def favicon_route():
    ico_bytes = generate_ico_bytes()
    if ico_bytes is None:
        return "", 404
    return send_file(
        BytesIO(ico_bytes),
        mimetype="image/x-icon",
    )


@app.route("/logo.png")
def logo_route():
    if not LOGO_BASE.exists():
        return "", 404
    return send_file(str(LOGO_BASE), mimetype="image/png")


def _markdown_to_html_simple(md_text: str) -> str:
    lines = md_text.splitlines()
    html_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            html_lines.append("<p></p>")
            continue
        if stripped.startswith("### "):
            html_lines.append(f"<h3>{stripped[4:]}</h3>")
        elif stripped.startswith("## "):
            html_lines.append(f"<h2>{stripped[3:]}</h2>")
        elif stripped.startswith("# "):
            html_lines.append(f"<h1>{stripped[2:]}</h1>")
        else:
            html_lines.append(f"<p>{stripped}</p>")
    return "\n".join(html_lines)


@app.route("/about")
def about_route():
    try:
        md_text = ABOUT_MD.read_text(encoding="utf-8")
    except Exception:
        md_text = "ABOUT.md kon niet worden gelezen."
    body_html = _markdown_to_html_simple(md_text)

    page = f"""<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <title>About CyNiT Cert Tool</title>
  <link rel="icon" type="image/x-icon" href="/favicon.ico">
  <style>
    body {{
      background: {BG_COLOR};
      color: {MAIN_COLOR};
      font-family: Arial, sans-serif;
      margin: 20px;
    }}
    h1, h2, h3 {{
      color: {TITLE_COLOR};
    }}
    a {{
      color: {MAIN_COLOR};
    }}
  </style>
</head>
<body>
{body_html}
</body>
</html>"""
    return page


# -------- Web profielbeheer ---------

@app.route("/profiles", methods=["GET"])
def profiles_page():
    cfg, active_name, base_profile = get_active_profile()
    return render_template_string(
        PROFILES_TEMPLATE,
        profiles=cfg["profiles"],
        active_profile=active_name,
        base_profile=base_profile,
    )


@app.route("/profiles/activate", methods=["POST"])
def profiles_activate():
    cfg, active_name, _ = get_active_profile()
    name = (request.form.get("profile_name") or "").strip()
    if not name or name not in cfg["profiles"]:
        return make_response("Profiel niet gevonden.", 400)
    if name == active_name:
        return redirect(url_for("profiles_page"))

    cfg["active_profile"] = name
    save_config(cfg)
    restart_program()  # auto-restart gehele app
    return ""  # wordt in praktijk niet bereikt


@app.route("/profiles/create", methods=["POST"])
def profiles_create():
    cfg, active_name, base_profile = get_active_profile()
    name = (request.form.get("name") or "").strip()
    if not name:
        return make_response("Profielnaam is verplicht.", 400)
    if name in cfg["profiles"]:
        return make_response("Profiel bestaat al.", 400)

    new_profile = copy.deepcopy(base_profile)

    def get_field(field, default):
        return request.form.get(field, default)

    colors = new_profile["colors"]
    paths = new_profile["paths"]
    ui = new_profile["ui"]

    colors["background"] = get_field("colors.background", colors["background"])
    colors["general_fg"] = get_field("colors.general_fg", colors["general_fg"])
    colors["title"] = get_field("colors.title", colors["title"])
    colors["table_col1_bg"] = get_field("colors.table_col1_bg", colors["table_col1_bg"])
    colors["table_col1_fg"] = get_field("colors.table_col1_fg", colors["table_col1_fg"])
    colors["table_col2_bg"] = get_field("colors.table_col2_bg", colors["table_col2_bg"])
    colors["table_col2_fg"] = get_field("colors.table_col2_fg", colors["table_col2_fg"])
    colors["button_bg"] = get_field("colors.button_bg", colors["button_bg"])
    colors["button_fg"] = get_field("colors.button_fg", colors["button_fg"])

    paths["logo"] = get_field("paths.logo", paths["logo"])
    paths["help"] = get_field("paths.help", paths["help"])

    logo_max = get_field("ui.logo_max_height", str(ui.get("logo_max_height", 80)))
    try:
        ui["logo_max_height"] = int(logo_max)
    except ValueError:
        ui["logo_max_height"] = ui.get("logo_max_height", 80)

    ui["font_main"] = get_field("ui.font_main", ui.get("font_main", "Consolas"))
    ui["font_buttons"] = get_field("ui.font_buttons", ui.get("font_buttons", "Segoe UI"))

    cfg["profiles"][name] = new_profile
    cfg["active_profile"] = name
    save_config(cfg)
    restart_program()
    return ""  # wordt niet bereikt


# ============================================================
#   ENTRYPOINTS
# ============================================================

def run_gui_only():
    gui = CynitCertGUI()
    gui.mainloop()


def run_web_only():
    app.run(host=WEB_HOST, port=WEB_PORT, debug=True)


def run_both():
    start_web_in_background()
    gui = CynitCertGUI()
    gui.mainloop()

if __name__ == "__main__":
    ensure_about_file()

    args = sys.argv[1:]

    # Standaard: alleen web
    if "--gui-only" in args:
        run_gui_only()
    elif "--both" in args:
        run_both()
    else:
        # geen args of expliciet --web => web-only
        run_web_only()
