#!/usr/bin/env python3
"""
cert_viewer.py

Certificate / CSR viewer module voor CyNiT Tools.

- Standalone web:
    python cert_viewer.py         -> http://127.0.0.1:5001/cert
- Standalone GUI:
    python cert_viewer.py --gui

- In CyNiT Tools hub (ctools.py):
    import cert_viewer
    settings = cynit_theme.load_settings()
    tools = cynit_theme.load_tools()["tools"]
    cert_viewer.register_web_routes(app, settings, tools)

Kleuren, logo en theming komen uit:
- cynit_theme.BASE_DIR / config/settings.json

Globale header/footer/wafel/hamburger komen uit:
- cynit_layout.py

Export-styling komt uit:
- config/exports.json (wordt automatisch aangemaakt met defaults).
"""

from __future__ import annotations

import sys
import os
import json
import csv
import subprocess
from pathlib import Path
from io import BytesIO
from zipfile import ZipFile
from typing import Dict, Any, List, Optional
from datetime import datetime

import tkinter as tk
from tkinter import filedialog, messagebox

from flask import (
    Flask,
    request,
    render_template_string,
    send_file,
    make_response,
)

from PIL import Image, ImageTk

from cryptography import x509
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import rsa, dsa, ec
from cryptography.x509.oid import NameOID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import cynit_theme
import cynit_layout
import cynit_exports


# ------------------------------------------------------------
#  Basis paden en config
# ------------------------------------------------------------

BASE_DIR: Path = cynit_theme.BASE_DIR
LOGO_PATH: Path = cynit_theme.LOGO_PATH
CONFIG_DIR: Path = cynit_theme.CONFIG_DIR
EXPORT_CONFIG_PATH: Path = CONFIG_DIR / "exports.json"

LAST_INFO: Optional[Dict[str, Any]] = None   # voor web-downloads

# Map waarin we vaste MD-exports bewaren
EXPORTS_DIR = BASE_DIR / "exports"


def ensure_exports_dir():
    """Zorgt dat de exports-map bestaat."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def slugify_filename(name: str) -> str:
    """
    Maakt een veilige bestandsnaam op basis van de originele filename.
    - Haalt de extensie eraf.
    - Vervangt vreemde tekens door '_'.
    """
    base = Path(name).stem or "certificate"
    base = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in base)
    return base or "certificate"


# ------------------------------------------------------------
#  Export-styles (xlsx / html / md) via exports.json
# ------------------------------------------------------------

def default_export_styles(settings: Dict[str, Any]) -> Dict[str, Any]:
    """
    Default stijlen voor exports.
    Kun je per profiel overschrijven door config/exports.json aan te passen.
    """
    return {
        "xlsx": {
            "sheet": {
                "default_bg": "#FFFFFF"
            },
            "title": {
                "font_size": 16,
                "bold": True,
                "italic": False,
                "color": "#000000"
            },
            "field_col": {
                "font_size": 12,
                "bold": True,
                "italic": True,
                "color": "#000000"
            },
            "value_col": {
                "font_size": 12,
                "bold": False,
                "italic": False,
                "color": "#000000"
            }
        },
        "html": {
            "body": {
                "bg": "#FFFFFF",
                "fg": "#000000"
            },
            "title": {
                "color": "#0000FF",
                "font_size_px": 16,
                "bold": True
            },
            "table": {
                "border_color": "#000000",
                "border_width_px": 1,
                "field_col": {
                    "font_size_px": 14,
                    "bold": True,
                    "italic": True
                },
                "value_col": {
                    "font_size_px": 12,
                    "bold": False,
                    "italic": False
                }
            }
        },
        "md": {
            "title_prefix": "# ",
            "section_prefix": "## ",
            "bold_field_names": True
        }
    }


def load_export_styles(settings: Dict[str, Any]) -> Dict[str, Any]:
    """
    Leest config/exports.json.
    - Bestaat de file niet -> maak hem met defaults.
    - Bestaat hij wel -> merge met defaults (zodat nieuwe keys blijven werken).
    """
    defaults = default_export_styles(settings)

    if not EXPORT_CONFIG_PATH.exists():
        EXPORT_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        EXPORT_CONFIG_PATH.write_text(json.dumps(defaults, indent=2), encoding="utf-8")
        return defaults

    try:
        current = json.loads(EXPORT_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        # Bij corrupte file: overschrijven met defaults
        EXPORT_CONFIG_PATH.write_text(json.dumps(defaults, indent=2), encoding="utf-8")
        return defaults

    try:
        merged = cynit_theme.deep_merge(defaults, current)
    except Exception:
        merged = defaults

    EXPORT_CONFIG_PATH.write_text(json.dumps(merged, indent=2), encoding="utf-8")
    return merged


# ------------------------------------------------------------
#  X.509 / CSR decode logica
# ------------------------------------------------------------

def load_cert_or_csr(data: bytes):
    """
    Probeert PEM/DER te detecteren als Certificate of CSR.

    Return:
        ("cert", x509.Certificate) of ("csr", x509.CertificateSigningRequest)
    Raise:
        ValueError bij mislukking.
    """
    text: Optional[str] = None
    try:
        text = data.decode("utf-8", errors="ignore")
    except Exception:
        pass

    # PEM CSR
    if text and "BEGIN CERTIFICATE REQUEST" in text:
        try:
            csr = x509.load_pem_x509_csr(data)
            return "csr", csr
        except Exception:
            pass

    # PEM CERT
    if text and "BEGIN CERTIFICATE" in text:
        try:
            cert = x509.load_pem_x509_certificate(data)
            return "cert", cert
        except Exception:
            pass

    # DER CERT
    try:
        cert = x509.load_der_x509_certificate(data)
        return "cert", cert
    except Exception:
        pass

    # DER CSR
    try:
        csr = x509.load_der_x509_csr(data)
        return "csr", csr
    except Exception:
        pass

    raise ValueError("Bestand is geen geldige X.509 certificate of CSR (PEM/DER).")


def get_name_attr(name: x509.Name, oid) -> str:
    try:
        attrs = name.get_attributes_for_oid(oid)
        if attrs:
            return attrs[0].value
    except Exception:
        pass
    return "-"


def subject_fields(name: x509.Name) -> Dict[str, str]:
    return {
        "Common Name":         get_name_attr(name, NameOID.COMMON_NAME),
        "emailAddress":        get_name_attr(name, NameOID.EMAIL_ADDRESS),
        "Organizational Unit": get_name_attr(name, NameOID.ORGANIZATIONAL_UNIT_NAME),
        "Organization":        get_name_attr(name, NameOID.ORGANIZATION_NAME),
        "Locality":            get_name_attr(name, NameOID.LOCALITY_NAME),
        "State or Province":   get_name_attr(name, NameOID.STATE_OR_PROVINCE_NAME),
        "Country":             get_name_attr(name, NameOID.COUNTRY_NAME),
    }


def issuer_fields(name: x509.Name) -> Dict[str, str]:
    return {
        "Issuer Common Name":       get_name_attr(name, NameOID.COMMON_NAME),
        "Issuer emailAddress":      get_name_attr(name, NameOID.EMAIL_ADDRESS),
        "Issuer Organization":      get_name_attr(name, NameOID.ORGANIZATION_NAME),
        "Issuer Locality":          get_name_attr(name, NameOID.LOCALITY_NAME),
        "Issuer State or Province": get_name_attr(name, NameOID.STATE_OR_PROVINCE_NAME),
        "Issuer Country":           get_name_attr(name, NameOID.COUNTRY_NAME),
    }


def format_name(name: x509.Name) -> str:
    parts = []
    for rdn in name.rdns:
        for attr in rdn:
            parts.append(f"{attr.oid._name}={attr.value}")
    return ", ".join(parts) if parts else "-"


def get_key_info(public_key):
    if isinstance(public_key, rsa.RSAPublicKey):
        return "RSA", str(public_key.key_size)
    if isinstance(public_key, dsa.DSAPublicKey):
        return "DSA", str(public_key.key_size)
    if isinstance(public_key, ec.EllipticCurvePublicKey):
        try:
            size = public_key.key_size
        except Exception:
            size = "-"
        return f"EC ({public_key.curve.name})", str(size)
    return public_key.__class__.__name__, "-"


def get_signature_algorithm(obj) -> str:
    try:
        sig_hash = obj.signature_hash_algorithm.name
    except Exception:
        sig_hash = "-"

    algo_name = "-"
    try:
        algo_name = obj.signature_algorithm_oid._name
    except Exception:
        pass

    if algo_name == "-":
        return sig_hash if sig_hash != "-" else "-"
    if sig_hash != "-":
        return f"{algo_name} ({sig_hash})"
    return algo_name


def compute_thumbprint(cert: x509.Certificate) -> str:
    try:
        fp = cert.fingerprint(hashes.SHA1())
        return fp.hex().upper()
    except Exception:
        return "-"


def get_validity_utc(obj):
    """
    Geeft (valid_from_iso, valid_to_iso) terug.
    Gebruikt *_utc als die bestaan (nieuwe cryptography),
    anders de oude not_valid_before / not_valid_after.
    """
    start = getattr(obj, "not_valid_before_utc", None)
    end = getattr(obj, "not_valid_after_utc", None)

    if start is None:
        start = obj.not_valid_before
    if end is None:
        end = obj.not_valid_after

    return start.isoformat(), end.isoformat()


def decode_cert_from_bytes(data: bytes, fake_path: Path) -> Dict[str, Any]:
    obj_type, obj = load_cert_or_csr(data)

    subj_map = subject_fields(obj.subject)

    if obj_type == "cert":
        issuer_map = issuer_fields(obj.issuer)
        valid_from, valid_to = get_validity_utc(obj)
        serial = hex(obj.serial_number).upper().replace("X", "x")
        thumb = compute_thumbprint(obj)
        issuer_str = format_name(obj.issuer)
    else:
        issuer_map = None
        valid_from = "-"
        valid_to = "-"
        serial = "-"
        thumb = "-"
        issuer_str = "-"

    pub = obj.public_key()
    key_algo, key_size = get_key_info(pub)
    sig_algo = get_signature_algorithm(obj)

    props = {
        "Subject":        format_name(obj.subject),
        "Issuer":         issuer_str,
        "Valid From":     valid_from,
        "Valid To":       valid_to,
        "Key Size":       key_size,
        "Key Algorithm":  key_algo,
        "Sig. Algorithm": sig_algo,
        "Serial Number":  serial,
        "Thumbprint":     thumb,
    }

    info: Dict[str, Any] = {
        "filename": str(fake_path),
        "type": "Certificate" if obj_type == "cert" else "CSR",
        "subject": subj_map,
        "issuer": issuer_map,
        "properties": props,
    }
    return info


def decode_cert_from_file(path: Path) -> Dict[str, Any]:
    return decode_cert_from_bytes(path.read_bytes(), path)


# ------------------------------------------------------------
#  Export builders (HTML / Markdown / XLSX / ZIP)
# ------------------------------------------------------------

def build_html_export(info: Dict[str, Any], settings: Dict[str, Any]) -> str:
    styles = load_export_styles(settings)
    hcfg = styles.get("html", {})

    body_cfg = hcfg.get("body", {})
    title_cfg = hcfg.get("title", {})
    table_cfg = hcfg.get("table", {})

    bg_body = body_cfg.get("bg", "#FFFFFF")
    fg_body = body_cfg.get("fg", "#000000")

    title_color = title_cfg.get("color", "#0000FF")
    title_size = title_cfg.get("font_size_px", 16)
    title_bold = title_cfg.get("bold", True)

    border_color = table_cfg.get("border_color", "#000000")
    border_width = table_cfg.get("border_width_px", 1)

    field_cfg = table_cfg.get("field_col", {})
    value_cfg = table_cfg.get("value_col", {})

    field_size = field_cfg.get("font_size_px", 14)
    field_bold = field_cfg.get("bold", True)
    field_italic = field_cfg.get("italic", True)

    value_size = value_cfg.get("font_size_px", 12)
    value_bold = value_cfg.get("bold", False)
    value_italic = value_cfg.get("italic", False)

    def font_style(bold: bool, italic: bool, size: int) -> str:
        weight = "bold" if bold else "normal"
        style = "italic" if italic else "normal"
        return f"font-weight: {weight}; font-style: {style}; font-size: {size}px;"

    field_style = font_style(field_bold, field_italic, field_size)
    value_style = font_style(value_bold, value_italic, value_size)
    title_weight = "bold" if title_bold else "normal"

    def dict_to_html_table(title_txt: str, mapping: Optional[Dict[str, Any]], issuer: bool = False) -> str:
        if issuer and mapping is None:
            return (
                f"<h2>{title_txt}</h2>"
                "<p>CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>"
            )
        if mapping is None:
            return ""
        rows = []
        for k, v in mapping.items():
            rows.append(
                "<tr>"
                f"<td style='{field_style}'>{k}</td>"
                f"<td style='{value_style}'>{v}</td>"
                "</tr>"
            )
        rows_html = "".join(rows)
        return f"<h2>{title_txt}</h2><table>{rows_html}</table>"

    subject_html = dict_to_html_table("Certificate Subject", info.get("subject"))
    issuer_html = dict_to_html_table("Certificate Issuer", info.get("issuer"), issuer=True)
    props_html = dict_to_html_table("Certificate Properties", info.get("properties"))

    html = f"""<!doctype html>
<html lang="nl">
<head>
<meta charset="utf-8">
<title>CyNiT Certificate Decoder Export</title>
<style>
body {{
  background: {bg_body};
  color: {fg_body};
  font-family: Arial, sans-serif;
}}
h1, h2 {{
  color: {title_color};
  font-size: {title_size}px;
  font-weight: {title_weight};
}}
table {{
  border-collapse: collapse;
  margin-bottom: 20px;
  min-width: 500px;
}}
td {{
  border: {border_width}px solid {border_color};
  padding: 4px 8px;
}}
</style>
</head>
<body>
<h1>CyNiT Certificate Decoder Export</h1>
<p>Bestand: {info.get("filename", "")}</p>
<p>Type: {info.get("type", "")}</p>
{subject_html}
{issuer_html}
{props_html}
</body>
</html>
"""
    return html


def build_markdown_export(info: Dict[str, Any], settings: Dict[str, Any]) -> str:
    styles = load_export_styles(settings)
    mcfg = styles.get("md", {})

    title_prefix = mcfg.get("title_prefix", "# ")
    section_prefix = mcfg.get("section_prefix", "## ")
    bold_field = mcfg.get("bold_field_names", True)

    def dict_to_md_table(title_txt: str, mapping: Optional[Dict[str, Any]], issuer: bool = False) -> str:
        if issuer and mapping is None:
            return f"{section_prefix}{title_txt}\n\nCSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.\n"
        if mapping is None:
            return ""
        lines = [f"{section_prefix}{title_txt}", "", "| Field | Value |", "| --- | --- |"]
        for k, v in mapping.items():
            field = f"**{k}**" if bold_field else k
            lines.append(f"| {field} | {v} |")
        lines.append("")
        return "\n".join(lines)

    md_parts: List[str] = []
    md_parts.append(f"{title_prefix}CyNiT Certificate Decoder Export\n")
    md_parts.append(f"**Bestand**: `{info.get('filename', '')}`  ")
    md_parts.append(f"**Type**: {info.get('type', '')}\n")

    md_parts.append(dict_to_md_table("Certificate Subject", info.get("subject")))
    md_parts.append(dict_to_md_table("Certificate Issuer", info.get("issuer"), issuer=True))
    md_parts.append(dict_to_md_table("Certificate Properties", info.get("properties")))

    return "\n".join(md_parts)


def build_xlsx_export(info: Dict[str, Any], settings: Dict[str, Any]) -> bytes:
    styles = load_export_styles(settings)
    xcfg = styles.get("xlsx", {})

    sheet_cfg = xcfg.get("sheet", {})
    title_cfg = xcfg.get("title", {})
    field_cfg = xcfg.get("field_col", {})
    value_cfg = xcfg.get("value_col", {})

    default_bg = sheet_cfg.get("default_bg", "#FFFFFF")

    title_font_size = title_cfg.get("font_size", 16)
    title_bold = title_cfg.get("bold", True)
    title_italic = title_cfg.get("italic", False)
    title_color = title_cfg.get("color", "#000000").lstrip("#")

    field_font_size = field_cfg.get("font_size", 12)
    field_bold = field_cfg.get("bold", True)
    field_italic = field_cfg.get("italic", True)
    field_color = field_cfg.get("color", "#000000").lstrip("#")

    value_font_size = value_cfg.get("font_size", 12)
    value_bold = value_cfg.get("bold", False)
    value_italic = value_cfg.get("italic", False)
    value_color = value_cfg.get("color", "#000000").lstrip("#")

    default_bg_hex = default_bg.lstrip("#")

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate"

    title_font = Font(
        color=title_color,
        bold=title_bold,
        italic=title_italic,
        size=title_font_size,
    )
    field_font = Font(
        color=field_color,
        bold=field_bold,
        italic=field_italic,
        size=field_font_size,
    )
    value_font = Font(
        color=value_color,
        bold=value_bold,
        italic=value_italic,
        size=value_font_size,
    )

    fill_default = PatternFill(
        start_color=default_bg_hex,
        end_color=default_bg_hex,
        fill_type="solid",
    )

    row = 1
    ws["A1"] = "CyNiT Certificate Decoder Export"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_default
    row += 2

    def set_row(key: str, value: str):
        nonlocal row
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = field_font
        ws[f"B{row}"].font = value_font
        ws[f"A{row}"].fill = fill_default
        ws[f"B{row}"].fill = fill_default
        row += 1

    set_row("Bestand", str(info.get("filename", "")))
    set_row("Type", str(info.get("type", "")))
    row += 1

    def write_section(title_txt: str, mapping: Optional[Dict[str, Any]]):
        nonlocal row
        ws[f"A{row}"] = title_txt
        ws[f"A{row}"].font = title_font
        ws[f"A{row}"].fill = fill_default
        row += 1
        if mapping is None:
            ws[f"A{row}"] = "CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat."
            ws[f"A{row}"].font = value_font
            ws[f"A{row}"].fill = fill_default
            row += 2
            return
        for k, v in mapping.items():
            ws[f"A{row}"] = k
            ws[f"B{row}"] = v
            ws[f"A{row}"].font = field_font
            ws[f"B{row}"].font = value_font
            ws[f"A{row}"].fill = fill_default
            ws[f"B{row}"].fill = fill_default
            row += 1
        row += 1

    write_section("Certificate Subject", info.get("subject"))
    write_section("Certificate Issuer", info.get("issuer"))
    write_section("Certificate Properties", info.get("properties"))

    # Kolombreedtes een beetje oprekken
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_zip_bytes(info: Dict[str, Any], settings: Dict[str, Any], formats: List[str]) -> bytes:
    """
    Bouwt een ZIP met de gevraagde formaten (json, csv, xlsx, html, md).
    Returnt bytes van het zip-bestand.
    """
    base_name = Path(info.get("filename", "certificate")).stem or "certificate"

    buf = BytesIO()
    with ZipFile(buf, "w") as z:
        for fmt in formats:
            fmt = fmt.lower()
            if fmt == "json":
                content = json.dumps(info, indent=2, ensure_ascii=False).encode("utf-8")
                z.writestr(f"{base_name}.json", content)
            elif fmt == "csv":
                lines = ["Section;Field;Value"]
                for section_key, section_name in [
                    ("subject", "Subject"),
                    ("issuer", "Issuer"),
                    ("properties", "Properties"),
                ]:
                    section = info.get(section_key)
                    if section is None:
                        continue
                    for k, v in section.items():
                        value_str = str(v).replace(";", ",")
                        lines.append(f"{section_name};{k};{value_str}")
                content = "\n".join(lines).encode("utf-8")
                z.writestr(f"{base_name}.csv", content)
            elif fmt == "html":
                html = build_html_export(info, settings)
                z.writestr(f"{base_name}.html", html.encode("utf-8"))
            elif fmt == "md":
                md = build_markdown_export(info, settings)
                z.writestr(f"{base_name}.md", md.encode("utf-8"))
            elif fmt == "xlsx":
                xlsx_bytes = build_xlsx_export(info, settings)
                z.writestr(f"{base_name}.xlsx", xlsx_bytes)
            else:
                continue

    buf.seek(0)
    return buf.getvalue()


# ------------------------------------------------------------
#  Simpele helpers voor web
# ------------------------------------------------------------

def set_last_info(info: Dict[str, Any]) -> None:
    global LAST_INFO
    LAST_INFO = info


def get_last_info() -> Optional[Dict[str, Any]]:
    return LAST_INFO


# ------------------------------------------------------------
#  Web-routes voor integratie in hub / standalone Flask
# ------------------------------------------------------------

def register_web_routes(app: Flask, settings: Dict[str, Any], tools=None) -> None:
    """
    Registreert /cert, /exports en alle download-routes in een bestaande Flask-app.

    Layout:
    - gebruikt cynit_layout.common_css() voor basis
    - gebruikt cynit_layout.header_html() + footer_html()
    - toont wafelmenu links (modules)
    - toont hamburger export-menu rechts op /cert als er 'info' is
    """
    colors = settings["colors"]

    BG = colors["background"]
    FG = colors["general_fg"]
    COL1_BG = colors["table_col1_bg"]
    COL1_FG = colors["table_col1_fg"]
    COL2_BG = colors["table_col2_bg"]
    COL2_FG = colors["table_col2_fg"]

    base_css = cynit_layout.common_css(settings)
    common_js = cynit_layout.common_js()

    extra_css = f"""
    .error {{
      color: #ff0000;
      font-weight: bold;
    }}
    table {{
      border-collapse: collapse;
      margin-bottom: 20px;
      min-width: 500px;
    }}
    th, td {{
      border: 1px solid #555;
      padding: 4px 8px;
    }}
    th {{
      background: {COL1_BG};
      color: {COL1_FG};
    }}
    td {{
      background: {COL2_BG};
      color: {COL2_FG};
    }}
    """

    # Hamburger rechts op /cert
    export_menu_html = """
      {% if info %}
      <div class="hamburger-wrapper">
        <div class="hamburger-icon" onclick="toggleExport()">☰</div>
        <div id="export-menu" class="hamburger-dropdown">
          <a href="/cert/download/json">⬇ JSON</a>
          <a href="/cert/download/csv">⬇ CSV</a>
          <a href="/cert/download/xlsx">⬇ XLSX</a>
          <a href="/cert/download/html">⬇ HTML</a>
          <a href="/cert/download/md">⬇ Markdown</a>
          <a href="/cert/download/zip_all">⬇ ZIP (alles)</a>
          <a href="/cert/zip_select">⬇ ZIP (selectie)</a>
          <a href="/cert/save_md">💾 Bewaar MD in exports/</a>
        </div>
      </div>
      {% endif %}
    """

    header_cert = cynit_layout.header_html(
        settings,
        tools=tools,
        title="CyNiT Certificate / CSR Viewer",
        right_html=export_menu_html,
    )
    footer = cynit_layout.footer_html()

    # -----------------------------
    # Hoofd-template voor /cert
    # -----------------------------
    def _build_main_template() -> str:
        additional_js = """
    function toggleExport() {
      var el = document.getElementById('export-menu');
      if (!el) return;
      el.style.display = (el.style.display === 'block') ? 'none' : 'block';
    }
        """
        template = (
            "<!doctype html>\n"
            "<html lang=\"nl\">\n"
            "<head>\n"
            "  <meta charset=\"utf-8\">\n"
            "  <title>CyNiT Certificate / CSR Viewer</title>\n"
            "  <link rel=\"icon\" type=\"image/x-icon\" href=\"/favicon.ico\">\n"
            "  <style>\n"
            + base_css
            + extra_css
            + "\n  </style>\n"
            "  <script>\n"
            + common_js
            + additional_js
            + "\n  </script>\n"
            "</head>\n"
            "<body>\n"
            + header_cert
            + "\n"
            "  <div class=\"page\">\n"
            "    <h1>Certificate / CSR Viewer</h1>\n"
            "\n"
            "    <form method=\"post\" enctype=\"multipart/form-data\">\n"
            "      <label>Upload certificaat of CSR:\n"
            "        <input type=\"file\" name=\"file\">\n"
            "      </label>\n"
            "      <button type=\"submit\">Decode</button>\n"
            "    </form>\n"
            "\n"
            "    {% if error %}<p class=\"error\">{{ error }}</p>{% endif %}\n"
            "\n"
            "    {% if info %}\n"
            "      <h2>Resultaat</h2>\n"
            "      <p><strong>Bestand:</strong> {{ info.filename }}</p>\n"
            "      <p><strong>Type:</strong> {{ info.type }}</p>\n"
            "\n"
            "      <h3>Certificate Subject</h3>\n"
            "      <table>\n"
            "        <tbody>\n"
            "          {% for k, v in info.subject.items() %}\n"
            "          <tr><th>{{ k }}</th><td>{{ v }}</td></tr>\n"
            "          {% endfor %}\n"
            "        </tbody>\n"
            "      </table>\n"
            "\n"
            "      <h3>Certificate Issuer</h3>\n"
            "      {% if info.issuer %}\n"
            "      <table>\n"
            "        <tbody>\n"
            "          {% for k, v in info.issuer.items() %}\n"
            "          <tr><th>{{ k }}</th><td>{{ v }}</td></tr>\n"
            "          {% endfor %}\n"
            "        </tbody>\n"
            "      </table>\n"
            "      {% else %}\n"
            "      <p>CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>\n"
            "      {% endif %}\n"
            "\n"
            "      <h3>Certificate Properties</h3>\n"
            "      <table>\n"
            "        <tbody>\n"
            "          {% for k, v in info.properties.items() %}\n"
            "          <tr><th>{{ k }}</th><td>{{ v }}</td></tr>\n"
            "          {% endfor %}\n"
            "        </tbody>\n"
            "      </table>\n"
            "    {% endif %}\n"
            "  </div>\n"
            "\n"
            + footer +
            "\n</body>\n</html>\n"
        )
        return template

    main_template = _build_main_template()

    @app.route("/cert", methods=["GET", "POST"])
    def cert_index():
        error = None
        info_obj = None

        if request.method == "POST":
            file = request.files.get("file")
            if not file or file.filename == "":
                error = "Geen bestand geselecteerd."
            else:
                try:
                    data = file.read()
                    info_obj = decode_cert_from_bytes(data, Path(file.filename))
                    set_last_info(info_obj)
                except Exception as e:
                    error = f"Fout bij decoderen: {e}"

        return render_template_string(
            main_template,
            error=error,
            info=info_obj,
            tools=tools,
        )

    # -----------------------------
    # Download-routes
    # -----------------------------
    @app.route("/cert/download/<fmt>", methods=["GET"])
    def cert_download(fmt: str):
        info = get_last_info()
        if info is None:
            return make_response("Nog geen certificaat/CSR gedecodeerd in deze sessie.", 400)

        base_name = Path(info.get("filename", "certificate")).stem or "certificate"

        if fmt == "json":
            content = json.dumps(info, indent=2, ensure_ascii=False)
            resp = make_response(content)
            resp.headers["Content-Type"] = "application/json; charset=utf-8"
            resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.json"'
            return resp

        if fmt == "csv":
            lines = ["Section;Field;Value"]
            for section_key, section_name in [
                ("subject", "Subject"),
                ("issuer", "Issuer"),
                ("properties", "Properties"),
            ]:
                section = info.get(section_key)
                if section is None:
                    continue
                for k, v in section.items():
                    value_str = str(v).replace(";", ",")
                    lines.append(f"{section_name};{k};{value_str}")
            content = "\n".join(lines)
            resp = make_response(content)
            resp.headers["Content-Type"] = "text/csv; charset=utf-8"
            resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.csv"'
            return resp

        if fmt == "html":
            html = build_html_export(info, settings)
            resp = make_response(html)
            resp.headers["Content-Type"] = "text/html; charset=utf-8"
            resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.html"'
            return resp

        if fmt == "md":
            md = build_markdown_export(info, settings)
            resp = make_response(md)
            resp.headers["Content-Type"] = "text/markdown; charset=utf-8"
            resp.headers["Content-Disposition"] = f'attachment; filename="{base_name}.md"'
            return resp

        if fmt == "xlsx":
            data = build_xlsx_export(info, settings)
            buf = BytesIO(data)
            buf.seek(0)
            return send_file(
                buf,
                as_attachment=True,
                download_name=f"{base_name}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        return make_response("Onbekend exporttype.", 400)

    @app.route("/cert/download/zip_all", methods=["GET"])
    def cert_zip_all():
        info = get_last_info()
        if info is None:
            return make_response("Nog geen certificaat/CSR gedecodeerd in deze sessie.", 400)

        formats = ["json", "csv", "xlsx", "html", "md"]
        zip_bytes = build_zip_bytes(info, settings, formats)
        base_name = Path(info.get("filename", "certificate")).stem or "certificate"
        buf = BytesIO(zip_bytes)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{base_name}_all.zip",
            mimetype="application/zip",
        )

    # -----------------------------
    # MD bewaren in exports/
    # -----------------------------
    @app.route("/cert/save_md", methods=["GET"])
    def cert_save_md():
        """
        Bewaart de huidige decode als Markdown in de map 'exports'
        en toont een korte bevestigingspagina.
        """
        info = get_last_info()
        if info is None:
            return make_response("Nog geen certificaat/CSR gedecodeerd in deze sessie.", 400)

        ensure_exports_dir()

        orig_name = info.get("filename", "certificate")
        slug = slugify_filename(orig_name)
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"{slug}_{ts}.md"
        dest = EXPORTS_DIR / filename

        md = build_markdown_export(info, settings)
        dest.write_text(md, encoding="utf-8")

        msg_html = f"""<!doctype html>
<html lang="nl">
<head>
  <meta charset="utf-8">
  <title>MD export opgeslagen</title>
  <link rel="icon" type="image/x-icon" href="/favicon.ico">
  <style>
    body {{
      background: {BG};
      color: {FG};
      font-family: Arial, sans-serif;
      margin: 20px;
    }}
    a {{
      color: {FG};
    }}
    code {{
      background: #222;
      padding: 2px 4px;
      border-radius: 3px;
    }}
  </style>
</head>
<body>
  <h1>MD export opgeslagen</h1>
  <p>De Markdown-export is bewaard als:</p>
  <p><code>exports/{filename}</code></p>
  <p>Je kan deze later openen in de Saved Exports pagina.</p>
  <p><a href="/cert">← Terug naar Cert Viewer</a></p>
</body>
</html>"""
        return msg_html

    # -----------------------------
    # ZIP selectie
    # -----------------------------
    @app.route("/cert/zip_select", methods=["GET", "POST"])
    def cert_zip_select():
        info = get_last_info()
        if info is None:
            return make_response("Nog geen certificaat/CSR gedecodeerd in deze sessie.", 400)

        all_formats = ["json", "csv", "xlsx", "html", "md"]

        if request.method == "POST":
            selected = request.form.getlist("fmt")
            selected = [f for f in selected if f in all_formats]
            if not selected:
                return make_response("Geen formaten geselecteerd.", 400)

            zip_bytes = build_zip_bytes(info, settings, selected)
            base_name = Path(info.get("filename", "certificate")).stem or "certificate"
            buf = BytesIO(zip_bytes)
            buf.seek(0)
            return send_file(
                buf,
                as_attachment=True,
                download_name=f"{base_name}_selected.zip",
                mimetype="application/zip",
            )

        base_css2 = cynit_layout.common_css(settings)
        common_js2 = cynit_layout.common_js()
        header2 = cynit_layout.header_html(
            settings,
            tools=tools,
            title="CyNiT Certificate / CSR Viewer",
            right_html="",
        )
        footer2 = cynit_layout.footer_html()

        form_html = (
            "<!doctype html>\n"
            "<html lang=\"nl\">\n"
            "<head>\n"
            "  <meta charset=\"utf-8\">\n"
            "  <title>Selecteer formaten - CyNiT Cert Viewer</title>\n"
            "  <link rel=\"icon\" type=\"image/x-icon\" href=\"/favicon.ico\">\n"
            "  <style>\n"
            + base_css2
            + "\n  </style>\n"
            "  <script>\n"
            + common_js2
            + "\n  </script>\n"
            "</head>\n"
            "<body>\n"
            + header2
            + "\n"
            "  <div class=\"page\">\n"
            "    <h1>Selecteer export-formaten</h1>\n"
            "    <p>Bestand: {{ filename }}</p>\n"
            "    <form method=\"post\">\n"
            "      <label><input type=\"checkbox\" name=\"fmt\" value=\"json\" checked> JSON</label><br>\n"
            "      <label><input type=\"checkbox\" name=\"fmt\" value=\"csv\" checked> CSV</label><br>\n"
            "      <label><input type=\"checkbox\" name=\"fmt\" value=\"xlsx\" checked> XLSX</label><br>\n"
            "      <label><input type=\"checkbox\" name=\"fmt\" value=\"html\" checked> HTML</label><br>\n"
            "      <label><input type=\"checkbox\" name=\"fmt\" value=\"md\" checked> Markdown</label><br><br>\n"
            "      <button type=\"submit\">Download ZIP</button>\n"
            "    </form>\n"
            "    <p><a href=\"/cert\">← Terug naar Cert Viewer</a></p>\n"
            "  </div>\n"
            "\n"
            + footer2 +
            "\n</body>\n</html>\n"
        )

        return render_template_string(
            form_html,
            filename=info.get("filename", ""),
            tools=tools,
        )

    # --------------------------------------------------------
    # Saved Exports pagina (/exports) + viewer (/exports/view)
    # --------------------------------------------------------
    ensure_exports_dir()

    header_exports = cynit_layout.header_html(
        settings,
        tools=tools,
        title="Saved Exports",
        right_html="",
    )

    exports_css = base_css  # zelfde look & feel

    exports_template = (
        "<!doctype html>\n"
        "<html lang=\"nl\">\n"
        "<head>\n"
        "  <meta charset=\"utf-8\">\n"
        "  <title>Saved Exports</title>\n"
        "  <link rel=\"icon\" type=\"image/x-icon\" href=\"/favicon.ico\">\n"
        "  <style>\n"
        + exports_css +
        "\n    table { border-collapse: collapse; width: 100%; }\n"
        "    th, td { border: 1px solid #333; padding: 4px 8px; }\n"
        "    th { text-align: left; }\n"
        "  </style>\n"
        "  <script>\n"
        + common_js +
        "\n  </script>\n"
        "</head>\n"
        "<body>\n"
        + header_exports +
        "\n"
        "  <div class=\"page\">\n"
        "    <h1>Saved Exports</h1>\n"
        "    <form method=\"get\" style=\"margin-bottom: 10px;\">\n"
        "      <label>Zoek: <input type=\"text\" name=\"q\" value=\"{{ query }}\" /></label>\n"
        "      <label style=\"margin-left:10px;\">Van (YYYY-MM-DD): <input type=\"text\" name=\"from\" value=\"{{ date_from }}\" size=\"10\"/></label>\n"
        "      <label style=\"margin-left:10px;\">Tot (YYYY-MM-DD): <input type=\"text\" name=\"to\" value=\"{{ date_to }}\" size=\"10\"/></label>\n"
        "      <button type=\"submit\">Filter</button>\n"
        "    </form>\n"
        "    {% if files %}\n"
        "    <table>\n"
        "      <thead><tr><th>Bestand</th><th>Titel</th><th>Laatste wijziging</th></tr></thead>\n"
        "      <tbody>\n"
        "        {% for f in files %}\n"
        "        <tr>\n"
        "          <td><a href=\"/exports/view/{{ f.name }}\">{{ f.name }}</a></td>\n"
        "          <td>{{ f.title }}</td>\n"
        "          <td>{{ f.mtime_str }}</td>\n"
        "        </tr>\n"
        "        {% endfor %}\n"
        "      </tbody>\n"
        "    </table>\n"
        "    {% else %}\n"
        "      <p>Er zijn nog geen exports gevonden in de map <code>exports/</code>.</p>\n"
        "    {% endif %}\n"
        "  </div>\n"
        "\n"
        + footer +
        "\n</body>\n</html>\n"
    )

    @app.route("/exports", methods=["GET"])
    @app.route("/exports/", methods=["GET"])
    def exports_index():
        from datetime import datetime, timedelta

        q = request.args.get("q", "").strip()
        date_from_str = request.args.get("from", "").strip()
        date_to_str = request.args.get("to", "").strip()

        dt_from = None
        dt_to = None

        def parse_date(val):
            try:
                return datetime.strptime(val, "%Y-%m-%d")
            except Exception:
                return None

        if date_from_str:
            dt_from = parse_date(date_from_str)
        if date_to_str:
            dt_to = parse_date(date_to_str)
            if dt_to:
                dt_to = dt_to + timedelta(days=1)  # inclusief einddag

        files_info = []
        for p in sorted(EXPORTS_DIR.glob("*.md"), key=lambda x: x.stat().st_mtime, reverse=True):
            stat = p.stat()
            mtime = datetime.fromtimestamp(stat.st_mtime)

            if dt_from and mtime < dt_from:
                continue
            if dt_to and mtime >= dt_to:
                continue

            text = p.read_text(encoding="utf-8", errors="ignore")
            if q:
                if q.lower() not in p.name.lower() and q.lower() not in text.lower():
                    continue

            title_line = ""
            for line in text.splitlines():
                if line.strip().startswith("#"):
                    title_line = line.lstrip("# ").strip()
                    break
            if not title_line:
                title_line = "(geen titel in MD)"

            files_info.append(
                {
                    "name": p.name,
                    "title": title_line,
                    "mtime_str": mtime.strftime("%Y-%m-%d %H:%M:%S"),
                }
            )

        return render_template_string(
            exports_template,
            files=files_info,
            query=q,
            date_from=date_from_str,
            date_to=date_to_str,
            tools=tools,
        )

    @app.route("/exports/view/<path:fname>", methods=["GET"])
    def exports_view(fname):
        # path traversal voorkomen
        safe_path = (EXPORTS_DIR / fname).resolve()
        if not safe_path.is_file() or safe_path.suffix.lower() != ".md":
            return make_response("Bestand niet gevonden.", 404)
        if EXPORTS_DIR.resolve() not in safe_path.parents:
            return make_response("Ongeldig pad.", 400)

        try:
            md = safe_path.read_text(encoding="utf-8")
        except Exception as e:
            return make_response(f"Kon bestand niet lezen: {e}", 500)

        body_html = cynit_theme.markdown_to_html_simple(md)

        page = (
            "<!doctype html>\n"
            "<html lang=\"nl\">\n"
            "<head>\n"
            "  <meta charset=\"utf-8\">\n"
            f"  <title>Export: {safe_path.name}</title>\n"
            "  <link rel=\"icon\" type=\"image/x-icon\" href=\"/favicon.ico\">\n"
            "  <style>\n"
            + base_css +
            "\n  </style>\n"
            "  <script>\n"
            + common_js +
            "\n  </script>\n"
            "</head>\n"
            "<body>\n"
            + header_exports +
            "\n"
            "  <div class=\"page\">\n"
            f"    <h1>{safe_path.name}</h1>\n"
            "    <div>\n"
            + body_html +
            "    </div>\n"
            "    <p><a href=\"/exports\">← Terug naar Saved Exports</a></p>\n"
            "  </div>\n"
            "\n"
            + footer +
            "\n</body>\n</html>\n"
        )
        return page

# ------------------------------------------------------------
#  GUI (Tkinter)
# ------------------------------------------------------------

class CertViewerGUI(tk.Tk):
    def __init__(self, settings: Dict[str, Any]):
        super().__init__()
        self.settings = settings
        colors = settings["colors"]
        ui = settings["ui"]

        self.title("CyNiT Certificate / CSR Viewer")
        self.geometry("1100x780")

        self.bg_color = colors["background"]
        self.fg_color = colors["general_fg"]
        self.title_color = colors["title"]

        self.col1_bg = colors["table_col1_bg"]
        self.col1_fg = colors["table_col1_fg"]
        self.col2_bg = colors["table_col2_bg"]
        self.col2_fg = colors["table_col2_fg"]

        self.button_bg = colors["button_bg"]
        self.button_fg = colors["button_fg"]

        self.font_main = ui.get("font_main", "Consolas")
        self.font_buttons = ui.get("font_buttons", "Segoe UI")
        self.logo_max_height = ui.get("logo_max_height", 80)

        self.base_font = (self.font_main, 12)
        self.button_font = (self.font_buttons, 11, "bold")
        self.label_font = (self.font_main, 11)

        self.configure(bg=self.bg_color)

        self.current_info: Optional[Dict[str, Any]] = None
        self.current_path: Optional[Path] = None
        self.export_buttons = []
        self.logo_img = None

        self._build_gui()

    def _build_gui(self) -> None:
        # Header (logo links)
        header = tk.Frame(self, bg=self.bg_color)
        header.pack(fill=tk.X, padx=10, pady=(10, 0))

        left = tk.Frame(header, bg=self.bg_color)
        left.pack(side=tk.LEFT, anchor="w")

        if LOGO_PATH.exists():
            try:
                img = Image.open(LOGO_PATH)
                if img.height > 0:
                    scale = self.logo_max_height / img.height
                else:
                    scale = 1.0
                new_w = int(img.width * scale)
                new_h = int(img.height * scale)
                img_resized = img.resize((new_w, new_h), Image.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(img_resized)
                logo_label = tk.Label(left, image=self.logo_img, bg=self.bg_color)
                logo_label.pack(side=tk.LEFT)
            except Exception:
                pass

        # Bovenbalk
        top_frame = tk.Frame(self, bg=self.bg_color)
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        left_frame = tk.Frame(top_frame, bg=self.bg_color)
        left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        btn_open = tk.Button(
            left_frame,
            text="Cert/CSR kiezen…",
            command=self.choose_file,
            bg=self.button_bg,
            fg=self.button_fg,
            activebackground=self.button_bg,
            activeforeground=self.button_fg,
            font=self.button_font,
            relief=tk.RAISED,
            bd=3,
        )
        btn_open.pack(side=tk.LEFT)

        self.lbl_file = tk.Label(
            left_frame,
            text="Geen bestand geselecteerd",
            bg=self.bg_color,
            fg=self.fg_color,
            anchor="w",
            font=self.label_font,
        )
        self.lbl_file.pack(side=tk.LEFT, padx=10)

        right_frame = tk.Frame(top_frame, bg=self.bg_color)
        right_frame.pack(side=tk.RIGHT, anchor="ne")

        btn_webui = tk.Button(
            right_frame,
            text="Open Web UI",
            command=self.open_web_ui,
            bg=self.button_bg,
            fg=self.button_fg,
            activebackground=self.button_bg,
            activeforeground=self.button_fg,
            font=self.button_font,
            width=20,
            relief=tk.RAISED,
            bd=3,
        )
        btn_webui.pack(side=tk.TOP, pady=2, anchor="e")

        # Export-knoppen
        def make_export_button(text: str, fmt: str):
            btn = tk.Button(
                right_frame,
                text=text,
                command=lambda f=fmt: self.export_current(f),
                bg=self.button_bg,
                fg=self.button_fg,
                activebackground=self.button_bg,
                activeforeground=self.button_fg,
                font=self.button_font,
                state=tk.DISABLED,
                width=20,
                relief=tk.RAISED,
                bd=3,
            )
            btn.pack(side=tk.TOP, pady=2, anchor="e")
            self.export_buttons.append(btn)

        make_export_button("Export JSON", "json")
        make_export_button("Export CSV", "csv")
        make_export_button("Export XLSX", "xlsx")
        make_export_button("Export HTML", "html")
        make_export_button("Export Markdown", "md")
        make_export_button("Export ALL", "all")

        # Tabel (scrollbaar)
        table_frame = tk.Frame(self, bg=self.bg_color)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.table_canvas = tk.Canvas(
            table_frame,
            bg=self.bg_color,
            highlightthickness=0,
        )
        self.table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.table_canvas.yview,
        )
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.table_canvas.configure(yscrollcommand=scrollbar.set)

        self.table_inner = tk.Frame(self.table_canvas, bg=self.bg_color)
        self.table_canvas.create_window((0, 0), window=self.table_inner, anchor="nw")

        def on_configure(event):
            self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all"))

        self.table_inner.bind("<Configure>", on_configure)

    def set_export_state(self, state) -> None:
        for btn in self.export_buttons:
            btn.config(state=state)

    def open_web_ui(self) -> None:
        url = "http://127.0.0.1:5001/cert"
        try:
            import webbrowser
            webbrowser.open(url)
        except Exception:
            messagebox.showinfo("Web UI", f"Open deze URL in je browser:\n{url}")

    def choose_file(self) -> None:
        filetypes = [
            ("Alle ondersteunde bestanden", "*.crt *.cer *.pem *.csr"),
            ("Certificates", "*.crt *.cer *.pem"),
            ("CSRs", "*.csr"),
            ("Alle bestanden", "*.*"),
        ]
        filename = filedialog.askopenfilename(
            title="Kies certificaat of CSR",
            filetypes=filetypes,
        )
        if not filename:
            return

        path = Path(filename)
        self.lbl_file.config(text=str(path))
        try:
            info = decode_cert_from_file(path)
        except Exception as e:
            messagebox.showerror("Fout", f"Kon bestand niet decoderen:\n{e}")
            self.set_export_state(tk.DISABLED)
            return

        self.current_info = info
        self.current_path = path
        self.set_export_state(tk.NORMAL)
        self.show_info(info)

    def clear_table(self) -> None:
        for w in self.table_inner.winfo_children():
            w.destroy()

    def show_info(self, info: Dict[str, Any]) -> None:
        self.clear_table()
        row = 0

        def section_title(txt: str):
            nonlocal row
            lbl = tk.Label(
                self.table_inner,
                text=txt,
                bg=self.bg_color,
                fg=self.title_color,
                font=(self.font_main, 13, "bold"),
            )
            lbl.grid(row=row, column=0, columnspan=2, sticky="w", pady=(10, 2))
            row += 1

        def separator():
            nonlocal row
            sep = tk.Label(
                self.table_inner,
                text="─" * 80,
                bg=self.bg_color,
                fg=self.fg_color,
                font=(self.font_main, 9),
            )
            sep.grid(row=row, column=0, columnspan=2, sticky="w", pady=(0, 5))
            row += 1

        def kv(key: str, value: Any):
            nonlocal row
            k_lbl = tk.Label(
                self.table_inner,
                text=key,
                bg=self.col1_bg,
                fg=self.col1_fg,
                font=self.label_font,
                anchor="w",
            )
            k_lbl.grid(row=row, column=0, sticky="nsew", padx=(0, 2), pady=1, ipadx=4, ipady=2)

            v_lbl = tk.Label(
                self.table_inner,
                text=str(value),
                bg=self.col2_bg,
                fg=self.col2_fg,
                font=self.label_font,
                anchor="w",
                justify="left",
                wraplength=800,
            )
            v_lbl.grid(row=row, column=1, sticky="nsew", padx=(2, 0), pady=1, ipadx=4, ipady=2)

            self.table_inner.grid_columnconfigure(0, weight=1)
            self.table_inner.grid_columnconfigure(1, weight=2)
            row += 1

        section_title(f"Bestand: {info.get('filename', '')}")
        kv("Type", info.get("type", ""))
        separator()

        section_title("Certificate Subject")
        for k, v in info.get("subject", {}).items():
            kv(k, v)
        separator()

        section_title("Certificate Issuer")
        if info.get("issuer") is None:
            kv("Issuer", "CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.")
        else:
            for k, v in info["issuer"].items():
                kv(k, v)
        separator()

        section_title("Certificate Properties")
        for k, v in info.get("properties", {}).items():
            kv(k, v)

    def export_current(self, fmt: str) -> None:
        if not self.current_info:
            messagebox.showwarning("Geen data", "Er is nog geen certificaat/CSR geladen.")
            return

        settings = self.settings

        if fmt == "all":
            base = filedialog.asksaveasfilename(
                title="Kies basenaam voor ALL export (zonder extensie)",
                defaultextension=".json",
                filetypes=[("JSON", "*.json"), ("Alle bestanden", "*.*")],
            )
            if not base:
                return
            base_path = Path(base).with_suffix("")
            try:
                # JSON
                base_name = base_path
                base_name.with_suffix(".json").write_text(
                    json.dumps(self.current_info, indent=2, ensure_ascii=False),
                    encoding="utf-8",
                )
                # CSV
                lines = ["Section;Field;Value"]
                for section_key, section_name in [
                    ("subject", "Subject"),
                    ("issuer", "Issuer"),
                    ("properties", "Properties"),
                ]:
                    section = self.current_info.get(section_key)
                    if section is None:
                        continue
                    for k, v in section.items():
                        value_str = str(v).replace(";", ",")
                        lines.append(f"{section_name};{k};{value_str}")
                base_name.with_suffix(".csv").write_text("\n".join(lines), encoding="utf-8")
                # XLSX
                xlsx_bytes = build_xlsx_export(self.current_info, settings)
                base_name.with_suffix(".xlsx").write_bytes(xlsx_bytes)
                # HTML
                html = build_html_export(self.current_info, settings)
                base_name.with_suffix(".html").write_text(html, encoding="utf-8")
                # MD
                md = build_markdown_export(self.current_info, settings)
                base_name.with_suffix(".md").write_text(md, encoding="utf-8")
            except Exception as e:
                messagebox.showerror("Export-fout", f"Export is mislukt:\n{e}")
                return

            messagebox.showinfo(
                "Export voltooid",
                "Alle formaten zijn aangemaakt:\n\n"
                f"{base_path.with_suffix('.json')}\n"
                f"{base_path.with_suffix('.csv')}\n"
                f"{base_path.with_suffix('.xlsx')}\n"
                f"{base_path.with_suffix('.html')}\n"
                f"{base_path.with_suffix('.md')}\n",
            )
            return

        ext_map = {
            "json": (".json", [("JSON", "*.json"), ("Alle bestanden", "*.*")]),
            "csv":  (".csv", [("CSV", "*.csv"), ("Alle bestanden", "*.*")]),
            "xlsx": (".xlsx", [("Excel XLSX", "*.xlsx"), ("Alle bestanden", "*.*")]),
            "html": (".html", [("HTML", "*.html"), ("Alle bestanden", "*.*")]),
            "md":   (".md", [("Markdown", "*.md"), ("Alle bestanden", "*.*")]),
        }

        default_ext, filetypes = ext_map[fmt]
        filename = filedialog.asksaveasfilename(
            title=f"Export {fmt.upper()} opslaan als",
            defaultextension=default_ext,
            filetypes=filetypes,
        )
        if not filename:
            return
        dest = Path(filename)
        try:
            if fmt == "json":
                dest.write_text(
                    json.dumps(self.current_info, indent=2, ensure_ascii=False),
                    encoding="utf-8",
                )
            elif fmt == "csv":
                lines = ["Section;Field;Value"]
                for section_key, section_name in [
                    ("subject", "Subject"),
                    ("issuer", "Issuer"),
                    ("properties", "Properties"),
                ]:
                    section = self.current_info.get(section_key)
                    if section is None:
                        continue
                    for k, v in section.items():
                        value_str = str(v).replace(";", ",")
                        lines.append(f"{section_name};{k};{value_str}")
                dest.write_text("\n".join(lines), encoding="utf-8")
            elif fmt == "xlsx":
                xlsx_bytes = build_xlsx_export(self.current_info, settings)
                dest.write_bytes(xlsx_bytes)
            elif fmt == "html":
                html = build_html_export(self.current_info, settings)
                dest.write_text(html, encoding="utf-8")
            elif fmt == "md":
                md = build_markdown_export(self.current_info, settings)
                dest.write_text(md, encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Export-fout", f"Export is mislukt:\n{e}")
            return
        messagebox.showinfo("Export voltooid", f"Export opgeslagen als:\n{dest}")


# ------------------------------------------------------------
#  Standalone entrypoints (web / gui)
# ------------------------------------------------------------

def restart_program() -> None:
    python = sys.executable
    args = sys.argv
    try:
        subprocess.Popen([python] + args, cwd=BASE_DIR)
    except Exception as e:
        print(f"[ERROR] Kon herstart niet uitvoeren: {e}")
    os._exit(0)


def run_gui() -> None:
    settings = cynit_theme.load_settings()
    gui = CertViewerGUI(settings)
    gui.mainloop()


def run_web() -> None:
    settings = cynit_theme.load_settings()
    app = Flask(__name__)
    register_web_routes(app, settings, tools=None)

    @app.route("/restart")
    def restart_route():
        restart_program()
        return ""  # wordt niet bereikt

    app.run(host="127.0.0.1", port=5001, debug=False)


if __name__ == "__main__":
    if "--gui" in sys.argv:
        run_gui()
    else:
        run_web()
