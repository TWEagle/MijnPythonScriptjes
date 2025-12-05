#!/usr/bin/env python3
"""
dcb_org_export.py

CyNiT Tools module:
- Leest per omgeving config uit config/dcbaas_api.json
- Haalt optioneel een access token uit env vars, dcbaas_api.json of plain-text token_file
- Roept /certificate/search aan per organisatie-code
- Bouwt een Excel met alle toepassingen + certificaten

Integratie in ctools.py:
    import dcb_org_export
    dcb_org_export.register_web_routes(app, SETTINGS, TOOLS)
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
from io import BytesIO
import os
import json
import datetime as dt

import requests
from flask import Flask, request, render_template_string, send_file

import cynit_theme
import cynit_layout
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Basis paden
BASE_DIR = cynit_theme.BASE_DIR
CONFIG_DIR = cynit_theme.CONFIG_DIR
DCBAAS_API_CFG = CONFIG_DIR / "dcbaas_api.json"


# ------------------------------------------------------------
#  Config / environment
# ------------------------------------------------------------

@dataclass
class EnvConfig:
    name: str
    label: str
    external_api_base: str
    access_token: Optional[str]
    token_file: Optional[str]


def _write_skeleton_dcbaas_api() -> None:
    """
    Maak een skeleton dcbaas_api.json aan als hij nog niet bestaat.
    Dit is enkel om iets te hebben dat je via /config-editor kunt invullen.
    """
    skeleton = {
        "default_env": "TI",
        "environments": {
            "DEV": {
                "label": "DCBaaS DEV",
                "external_api_base": "",
                "access_token": "",
                "token_file": ""
            },
            "TI": {
                "label": "DCBaaS TI",
                "external_api_base": "",
                "access_token": "",
                "token_file": ""
            },
            "PROD": {
                "label": "DCBaaS PROD",
                "external_api_base": "",
                "access_token": "",
                "token_file": ""
            }
        }
    }
    DCBAAS_API_CFG.parent.mkdir(parents=True, exist_ok=True)
    DCBAAS_API_CFG.write_text(json.dumps(skeleton, indent=2), encoding="utf-8")


def load_env_configs_from_dcbaas_api() -> Tuple[Dict[str, EnvConfig], Optional[str]]:
    """
    Leest config/dcbaas_api.json en bouwt EnvConfig per environment.

    Verwachte structuur:

    {
      "default_env": "TI",
      "environments": {
        "DEV": {
          "label": "DCBaaS DEV",
          "external_api_base": "https://extapi.dcb-dev...",
          "access_token": "",
          "token_file": "C:/.../API/dev/access_token.txt"
        },
        ...
      }
    }
    """
    if not DCBAAS_API_CFG.exists():
        _write_skeleton_dcbaas_api()

    try:
        raw = json.loads(DCBAAS_API_CFG.read_text(encoding="utf-8"))
    except Exception:
        _write_skeleton_dcbaas_api()
        raw = json.loads(DCBAAS_API_CFG.read_text(encoding="utf-8"))

    envs_raw = raw.get("environments", {})
    envs: Dict[str, EnvConfig] = {}

    for env_key, cfg in envs_raw.items():
        if not isinstance(cfg, dict):
            continue
        envs[env_key] = EnvConfig(
            name=env_key,
            label=cfg.get("label", env_key),
            external_api_base=cfg.get("external_api_base", "") or "",
            access_token=(cfg.get("access_token") or None),
            token_file=(cfg.get("token_file") or None),
        )

    default_env = raw.get("default_env")
    if not isinstance(default_env, str):
        default_env = None

    # Fallback als environments leeg is
    if not envs:
        envs["DEV"] = EnvConfig(
            name="DEV",
            label="DCBaaS DEV",
            external_api_base="",
            access_token=None,
            token_file=None,
        )

    return envs, default_env


def load_token_from_file(token_file: str | None) -> str:
    """
    Leest een token uit token_file als plain text.

    Verwacht een bestand dat enkel de token-string bevat,
    b.v. 'Bearer eyJ...'
    """
    if not token_file:
        return ""

    p = Path(token_file)
    if not p.exists():
        return ""

    try:
        raw = p.read_text(encoding="utf-8").strip()
    except Exception:
        return ""

    return raw or ""


def load_default_token_for_env(env: EnvConfig) -> str:
    """
    Probeert een default access token op te pikken voor een omgeving:

    1) Omgevingsvariabele DCBAAS_TOKEN_<ENV>
    2) Omgevingsvariabele DCBAAS_TOKEN
    3) 'access_token' uit dcbaas_api.json (als ingevuld)
    4) 'token_file' uit dcbaas_api.json (plain-text bestand)
    """
    env_specific = os.getenv(f"DCBAAS_TOKEN_{env.name.upper()}")
    if env_specific:
        return env_specific.strip()

    generic = os.getenv("DCBAAS_TOKEN")
    if generic:
        return generic.strip()

    if env.access_token and env.access_token.strip():
        return env.access_token.strip()

    return load_token_from_file(env.token_file)


# ------------------------------------------------------------
#  API-call naar /certificate/search
# ------------------------------------------------------------

def build_certificate_search_body(org_code: str) -> Dict[str, Any]:
    """
    Bouwt de body voor /certificate/search.

    LET OP:
    Pas deze structuur aan indien de API een andere filter verwacht.
    Voor nu gaan we uit van een filter 'organization_code'.
    """
    return {
        "organization_code": org_code
    }


def fetch_certificates_for_org(
    env: EnvConfig,
    org_code: str,
    access_token: str,
    timeout: int = 30,
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    """
    Roept /certificate/search aan voor één organisatie-code.

    access_token = exacte string die in de Authorization-header moet,
    bv. "Bearer eyJ...".
    """
    if not env.external_api_base:
        return [], f"Base URL voor omgeving {env.name} is nog niet ingevuld in dcbaas_api.json."

    url = env.external_api_base.rstrip("/") + "/certificate/search"
    body = build_certificate_search_body(org_code)

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        # Token raw, net zoals in je bestaande certificates_gui
        "Authorization": access_token.strip(),
    }

    try:
        resp = requests.post(url, json=body, headers=headers, timeout=timeout)
    except Exception as exc:
        return [], f"HTTP-fout voor {org_code}: {exc}"

    if resp.status_code != 200:
        msg = resp.text
        if len(msg) > 300:
            msg = msg[:300] + "..."
        return [], f"Status {resp.status_code} bij {url}: {msg}"

    try:
        data = resp.json()
    except Exception as exc:
        return [], f"Kon JSON niet parsen voor {org_code}: {exc}"

    items = data.get("response")
    if not isinstance(items, list):
        if isinstance(data, list):
            items = data
        else:
            return [], f"Onverwacht antwoord voor {org_code}: geen 'response' lijst."

    return items, None


# ------------------------------------------------------------
#  Excel export
# ------------------------------------------------------------

def build_excel(results: Dict[str, List[Dict[str, Any]]]) -> bytes:
    """
    Maakt één XLSX met alle organisaties.

    Sheet 'Certificates':
      - organization_code
      - application_name
      - application_status
      - contact_persons
      - description
      - type
      - issued_by
      - start_date
      - end_date
      - status
      - serial_number

    Sheet 'Applications':
      Unieke toepassingen per organisatie.
    """
    wb = Workbook()
    ws_cert = wb.active
    ws_cert.title = "Certificates"

    cert_headers = [
        "organization_code",
        "application_name",
        "application_status",
        "contact_persons",
        "description",
        "type",
        "issued_by",
        "start_date",
        "end_date",
        "status",
        "serial_number",
    ]
    ws_cert.append(cert_headers)

    # Certificates sheet vullen
    for org_code, items in results.items():
        for row in items:
            contact = row.get("contact_person") or row.get("contact_persons")
            if isinstance(contact, list):
                contact_str = ", ".join(str(c) for c in contact)
            else:
                contact_str = str(contact) if contact is not None else ""

            ws_cert.append([
                org_code,
                row.get("application_name", ""),
                row.get("application_status", ""),
                contact_str,
                row.get("description", ""),
                row.get("type", ""),
                row.get("issued_by", ""),
                row.get("start_date", ""),
                row.get("end_date", ""),
                row.get("status", ""),
                row.get("serial_number", ""),
            ])

    # Applications sheet (unieke toepassingen per org)
    ws_app = wb.create_sheet("Applications")
    app_headers = [
        "organization_code",
        "application_name",
        "application_status",
        "contact_persons",
        "description",
        "type",
    ]
    ws_app.append(app_headers)

    seen = set()
    for org_code, items in results.items():
        for row in items:
            app_name = row.get("application_name", "")
            key = (org_code, app_name)
            if key in seen:
                continue
            seen.add(key)

            contact = row.get("contact_person") or row.get("contact_persons")
            if isinstance(contact, list):
                contact_str = ", ".join(str(c) for c in contact)
            else:
                contact_str = str(contact) if contact is not None else ""

            ws_app.append([
                org_code,
                app_name,
                row.get("application_status", ""),
                contact_str,
                row.get("description", ""),
                row.get("type", ""),
            ])

    # Kolombreedte automatisch
    for ws in (ws_cert, ws_app):
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws[f"{col_letter}{row_idx}"].value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ------------------------------------------------------------
#  Web UI
# ------------------------------------------------------------

def register_web_routes(app: Flask, settings: Dict[str, Any], tools=None) -> None:
    """
    Integreer deze tool in de bestaande CyNiT Tools Flask-app.

    Route:
      - GET/POST /dcbaas-org-export
    """
    envs, default_env = load_env_configs_from_dcbaas_api()
    base_css = cynit_layout.common_css(settings)
    common_js = cynit_layout.common_js()

    colors = settings.get("colors", {})
    bg = colors.get("background", "#000000")
    fg = colors.get("general_fg", "#FFFFFF")
    title_color = colors.get("title", "#00A2FF")
    t1_bg = colors.get("table_col1_bg", "#333333")
    t1_fg = colors.get("table_col1_fg", "#000000")
    t2_bg = colors.get("table_col2_bg", "#111111")
    t2_fg = colors.get("table_col2_fg", "#00FA00")
    btn_bg = colors.get("button_bg", "#111111")
    btn_fg = colors.get("button_fg", "#00B7C3")

    header = cynit_layout.header_html(
        settings,
        tools=tools,
        title="DCBaaS – Export per organisatie",
        right_html="",
    )
    footer = cynit_layout.footer_html()

    extra_css = f"""
    .card {{
      max-width: 1100px;
      margin: 0 auto 20px auto;
      background: #111111;
      padding: 20px;
      border-radius: 16px;
      box-shadow: 0 10px 30px rgba(0,0,0,0.7);
      color: {fg};
    }}
    h1, h2 {{
      color: {title_color};
      margin-top: 0;
    }}
    label {{
      display:block;
      margin-top:12px;
      font-weight:600;
    }}
    textarea, select, input[type="text"] {{
      width:100%;
      padding:8px 10px;
      border-radius:8px;
      border:1px solid #444;
      background:{bg};
      color:{fg};
      box-sizing:border-box;
    }}
    textarea {{
      min-height:120px;
      font-family:Consolas, monospace;
    }}
    .btn {{
      margin-top:16px;
      padding:8px 16px;
      border-radius:999px;
      border:1px solid #333;
      background:{btn_bg};
      color:{btn_fg};
      font-weight:700;
      cursor:pointer;
      display:inline-block;
      margin-right:10px;
    }}
    .btn:hover {{
      filter:brightness(1.15);
    }}
    .muted {{
      color:#aaa;
      font-size:0.9em;
    }}
    .error {{
      color:#fecaca;
      background:#7f1d1d;
      padding:8px 12px;
      border-radius:8px;
      margin-bottom:10px;
    }}
    table {{
      border-collapse: collapse;
      width: 100%;
      margin-top: 10px;
      font-size:0.9em;
    }}
    th, td {{
      border: 1px solid #333;
      padding: 4px 6px;
    }}
    th {{
      background: {t1_bg};
      color: {t1_fg};
    }}
    tbody tr:nth-child(odd) {{
      background: {t2_bg};
      color: {t2_fg};
    }}
    tbody tr:nth-child(even) {{
      background: #050505;
      color: {fg};
    }}
    """

    page_template = (
        "<!doctype html>\n"
        "<html lang='nl'>\n"
        "<head>\n"
        "  <meta charset='utf-8'>\n"
        "  <title>DCBaaS – Export per organisatie</title>\n"
        "  <style>\n"
        f"{base_css}\n{extra_css}\n"
        "  </style>\n"
        "  <script>\n"
        f"{common_js}\n"
        "  </script>\n"
        "</head>\n"
        "<body>\n"
        f"{header}\n"
        "<div class='page'>\n"
        "  <div class='card'>\n"
        "    <h1>DCBaaS – Export per organisatie</h1>\n"
        "    <p class='muted'>\n"
        "      1. Genereer eerst een access token via je bestaande JWT/JWK tooling.<br>\n"
        "      2. Plak hieronder exact wat je ook in je Certificates GUI bij 'Access token' zet\n"
        "         (bv. <code>Bearer eyJ...</code>).<br>\n"
        "      3. Vul één of meerdere organisatie-codes in (één per lijn) en kies Preview of Excel.\n"
        "    </p>\n"
        "    {% if error %}\n"
        "      <div class='error'>{{ error }}</div>\n"
        "    {% endif %}\n"
        "    <form method='post'>\n"
        "      <label>Omgeving</label>\n"
        "      <select name='env'>\n"
        "        {% for key, env in envs.items() %}\n"
        "          <option value='{{ key }}' {% if key == current_env %}selected{% endif %}>\n"
        "            {{ key }} – {{ env.label }} ({{ env.external_api_base }})\n"
        "          </option>\n"
        "        {% endfor %}\n"
        "      </select>\n"
        "      <label>Access token (Authorization header)</label>\n"
        "      <input type='text' name='access_token' value='{{ access_token }}' />\n"
        "      <p class='muted'>Bijvoorbeeld: <code>Bearer eyJ...</code>. Laat dit niet leeg.</p>\n"
        "      <label>Organisatie-codes</label>\n"
        "      <textarea name='org_codes' "
        "placeholder='OVO000082&#10;OVO002949'>{{ org_input }}</textarea>\n"
        "      <p class='muted'>Lege lijnen worden genegeerd. Copy/paste uit Excel mag.</p>\n"
        "      <button type='submit' name='action' value='preview' class='btn'>Voorbeeld tonen</button>\n"
        "      <button type='submit' name='action' value='export' class='btn'>Excel downloaden</button>\n"
        "    </form>\n"
        "  </div>\n"
        "  {% if preview %}\n"
        "    <div class='card'>\n"
        "      <h2>Preview resultaten</h2>\n"
        "      {% if total == 0 %}\n"
        "        <p class='muted'>Geen certificaten gevonden voor de opgegeven codes.</p>\n"
        "      {% else %}\n"
        "        <p class='muted'>Totaal {{ total }} certificaten voor {{ org_count }} organisaties.</p>\n"
        "        <table>\n"
        "          <thead>\n"
        "            <tr>\n"
        "              <th>Org</th><th>Application</th><th>App status</th>\n"
        "              <th>Cert status</th><th>Serial</th><th>Start</th><th>End</th>\n"
        "            </tr>\n"
        "          </thead>\n"
        "          <tbody>\n"
        "            {% for row in preview_rows %}\n"
        "            <tr>\n"
        "              <td>{{ row.org }}</td>\n"
        "              <td>{{ row.app }}</td>\n"
        "              <td>{{ row.app_status }}</td>\n"
        "              <td>{{ row.cert_status }}</td>\n"
        "              <td>{{ row.serial }}</td>\n"
        "              <td>{{ row.start }}</td>\n"
        "              <td>{{ row.end }}</td>\n"
        "            </tr>\n"
        "            {% endfor %}\n"
        "          </tbody>\n"
        "        </table>\n"
        "      {% endif %}\n"
        "      {% if errors %}\n"
        "        <h3>Fouten / waarschuwingen</h3>\n"
        "        <ul>\n"
        "          {% for e in errors %}<li>{{ e }}</li>{% endfor %}\n"
        "        </ul>\n"
        "      {% endif %}\n"
        "    </div>\n"
        "  {% endif %}\n"
        "</div>\n"
        f"{footer}\n"
        "</body>\n"
        "</html>\n"
    )

    def _render(**ctx):
        return render_template_string(page_template, tools=tools, **ctx)

    # Default environment bepalen
    if default_env and default_env in envs:
        initial_env = default_env
    elif "DEV" in envs:
        initial_env = "DEV"
    else:
        initial_env = next(iter(envs.keys()))

    @app.route("/dcbaas-org-export", methods=["GET", "POST"])
    def dcbaas_org_export():
        error: Optional[str] = None
        org_input = ""
        access_token = ""
        preview = False
        preview_rows: List[Dict[str, Any]] = []
        errors: List[str] = []
        total = 0

        current_env_key = initial_env

        if request.method == "GET":
            env = envs.get(current_env_key, next(iter(envs.values())))
            access_token_local = load_default_token_for_env(env)
            return _render(
                envs=envs,
                current_env=current_env_key,
                error=None,
                org_input="",
                access_token=access_token_local,
                preview=False,
                preview_rows=[],
                total=0,
                org_count=0,
                errors=[],
            )

        # POST
        current_env_key = request.form.get("env", current_env_key)
        env = envs.get(current_env_key, next(iter(envs.values())))

        org_input = request.form.get("org_codes", "") or ""
        access_token = request.form.get("access_token", "") or ""
        action = request.form.get("action") or "preview"

        org_codes = [line.strip() for line in org_input.splitlines() if line.strip()]

        if not access_token.strip():
            error = "Geef een access token in (Authorization header waarde)."
        elif not org_codes:
            error = "Geef minstens één organisatie-code in."

        results_by_org: Dict[str, List[Dict[str, Any]]] = {}

        if not error:
            for org in org_codes:
                items, err = fetch_certificates_for_org(env, org, access_token)
                if err:
                    errors.append(err)
                results_by_org[org] = items
                total += len(items)

            if action == "export":
                xlsx_bytes = build_excel(results_by_org)
                buf = BytesIO(xlsx_bytes)
                buf.seek(0)
                ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
                filename = f"dcbaas_org_export_{current_env_key}_{ts}.xlsx"
                return send_file(
                    buf,
                    as_attachment=True,
                    download_name=filename,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # Preview
            preview = True
            max_rows = 50
            for org, items in results_by_org.items():
                for row in items:
                    preview_rows.append({
                        "org": org,
                        "app": row.get("application_name", ""),
                        "app_status": row.get("application_status", ""),
                        "cert_status": row.get("status", ""),
                        "serial": row.get("serial_number", ""),
                        "start": row.get("start_date", ""),
                        "end": row.get("end_date", ""),
                    })
                    if len(preview_rows) >= max_rows:
                        break
                if len(preview_rows) >= max_rows:
                    break

        org_count = len({r["org"] for r in preview_rows}) if preview_rows else 0

        return _render(
            envs=envs,
            current_env=current_env_key,
            error=error,
            org_input=org_input,
            access_token=access_token,
            preview=preview,
            preview_rows=preview_rows,
            total=total,
            org_count=org_count,
            errors=errors,
        )


# Standalone web-run (optioneel)
if __name__ == "__main__":
    settings = cynit_theme.load_settings()
    tools_cfg = cynit_theme.load_tools()
    tools = tools_cfg.get("tools", [])
    app = Flask(__name__)
    register_web_routes(app, settings, tools)
    app.run(host="127.0.0.1", port=5451, debug=True)
