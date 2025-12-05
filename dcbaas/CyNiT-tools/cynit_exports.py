#!/usr/bin/env python3
"""
cynit_exports.py

Gedeelde export-logica (HTML / Markdown / XLSX / ZIP) voor CyNiT Tools.

Wordt gebruikt door o.a.:
- cert_viewer.py
en kan later hergebruikt worden door andere tools.

Functies die je normaal wil gebruiken:

- ensure_exports_dir()       → zorgt dat BASE_DIR / "exports" bestaat
- slugify_filename(name)     → veilige, nette bestandsnaam
- load_export_styles(settings)
- build_html_export(info, settings)
- build_markdown_export(info, settings)
- build_xlsx_export(info, settings)
- build_zip_bytes(info, settings, formats)

En constanten:

- BASE_DIR
- CONFIG_DIR
- EXPORT_CONFIG_PATH
- EXPORTS_DIR
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, List, Optional
from io import BytesIO
from zipfile import ZipFile
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import cynit_theme


# ------------------------------------------------------------
#  Basis paden en config
# ------------------------------------------------------------

BASE_DIR: Path = cynit_theme.BASE_DIR
CONFIG_DIR: Path = cynit_theme.CONFIG_DIR
EXPORT_CONFIG_PATH: Path = CONFIG_DIR / "exports.json"

# Map waarin we vaste MD-exports bewaren (kan gedeeld worden)
EXPORTS_DIR: Path = BASE_DIR / "exports"


def ensure_exports_dir() -> None:
    """Zorgt dat de exports-map (BASE_DIR/exports) bestaat."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def slugify_filename(name: str) -> str:
    """
    Maakt een veilige bestandsnaam op basis van de originele filename.
    - Haalt de extensie eraf.
    - Vervangt vreemde tekens door '_'.
    """
    base = Path(name).stem or "certificate"
    base = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in base)
    return base or "certificate"


# ------------------------------------------------------------
#  Export-styles (xlsx / html / md) via exports.json
# ------------------------------------------------------------

def default_export_styles(settings: Dict[str, Any]) -> Dict[str, Any]:
    """
    Default stijlen voor exports.
    Kun je per profiel overschrijven door config/exports.json aan te passen.
    """
    return {
        "xlsx": {
            "sheet": {
                "default_bg": "#FFFFFF"
            },
            "title": {
                "font_size": 16,
                "bold": True,
                "italic": False,
                "color": "#000000"
            },
            "field_col": {
                "font_size": 12,
                "bold": True,
                "italic": True,
                "color": "#000000"
            },
            "value_col": {
                "font_size": 12,
                "bold": False,
                "italic": False,
                "color": "#000000"
            }
        },
        "html": {
            "body": {
                "bg": "#FFFFFF",
                "fg": "#000000"
            },
            "title": {
                "color": "#0000FF",
                "font_size_px": 16,
                "bold": True
            },
            "table": {
                "border_color": "#000000",
                "border_width_px": 1,
                "field_col": {
                    "font_size_px": 14,
                    "bold": True,
                    "italic": True
                },
                "value_col": {
                    "font_size_px": 12,
                    "bold": False,
                    "italic": False
                }
            }
        },
        "md": {
            "title_prefix": "# ",
            "section_prefix": "## ",
            "bold_field_names": True
        }
    }


def load_export_styles(settings: Dict[str, Any]) -> Dict[str, Any]:
    """
    Leest config/exports.json.
    - Bestaat de file niet -> maak hem met defaults.
    - Bestaat hij wel -> merge met defaults (zodat nieuwe keys blijven werken).
    """
    defaults = default_export_styles(settings)

    if not EXPORT_CONFIG_PATH.exists():
        EXPORT_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        EXPORT_CONFIG_PATH.write_text(json.dumps(defaults, indent=2), encoding="utf-8")
        return defaults

    try:
        current = json.loads(EXPORT_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        # Bij corrupte file: overschrijven met defaults
        EXPORT_CONFIG_PATH.write_text(json.dumps(defaults, indent=2), encoding="utf-8")
        return defaults

    try:
        merged = cynit_theme.deep_merge(defaults, current)
    except Exception:
        merged = defaults

    EXPORT_CONFIG_PATH.write_text(json.dumps(merged, indent=2), encoding="utf-8")
    return merged


# ------------------------------------------------------------
#  Export builders (HTML / Markdown / XLSX / ZIP)
# ------------------------------------------------------------

def build_html_export(info: Dict[str, Any], settings: Dict[str, Any]) -> str:
    styles = load_export_styles(settings)
    hcfg = styles.get("html", {})

    body_cfg = hcfg.get("body", {})
    title_cfg = hcfg.get("title", {})
    table_cfg = hcfg.get("table", {})

    bg_body = body_cfg.get("bg", "#FFFFFF")
    fg_body = body_cfg.get("fg", "#000000")

    title_color = title_cfg.get("color", "#0000FF")
    title_size = title_cfg.get("font_size_px", 16)
    title_bold = title_cfg.get("bold", True)

    border_color = table_cfg.get("border_color", "#000000")
    border_width = table_cfg.get("border_width_px", 1)

    field_cfg = table_cfg.get("field_col", {})
    value_cfg = table_cfg.get("value_col", {})

    field_size = field_cfg.get("font_size_px", 14)
    field_bold = field_cfg.get("bold", True)
    field_italic = field_cfg.get("italic", True)

    value_size = value_cfg.get("font_size_px", 12)
    value_bold = value_cfg.get("bold", False)
    value_italic = value_cfg.get("italic", False)

    def font_style(bold: bool, italic: bool, size: int) -> str:
        weight = "bold" if bold else "normal"
        style = "italic" if italic else "normal"
        return f"font-weight: {weight}; font-style: {style}; font-size: {size}px;"

    field_style = font_style(field_bold, field_italic, field_size)
    value_style = font_style(value_bold, value_italic, value_size)
    title_weight = "bold" if title_bold else "normal"

    def dict_to_html_table(title_txt: str, mapping: Optional[Dict[str, Any]], issuer: bool = False) -> str:
        if issuer and mapping is None:
            return (
                f"<h2>{title_txt}</h2>"
                "<p>CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.</p>"
            )
        if mapping is None:
            return ""
        rows = []
        for k, v in mapping.items():
            rows.append(
                "<tr>"
                f"<td style='{field_style}'>{k}</td>"
                f"<td style='{value_style}'>{v}</td>"
                "</tr>"
            )
        rows_html = "".join(rows)
        return f"<h2>{title_txt}</h2><table>{rows_html}</table>"

    subject_html = dict_to_html_table("Certificate Subject", info.get("subject"))
    issuer_html = dict_to_html_table("Certificate Issuer", info.get("issuer"), issuer=True)
    props_html = dict_to_html_table("Certificate Properties", info.get("properties"))

    html = f"""<!doctype html>
<html lang="nl">
<head>
<meta charset="utf-8">
<title>CyNiT Certificate Decoder Export</title>
<style>
body {{
  background: {bg_body};
  color: {fg_body};
  font-family: Arial, sans-serif;
}}
h1, h2 {{
  color: {title_color};
  font-size: {title_size}px;
  font-weight: {title_weight};
}}
table {{
  border-collapse: collapse;
  margin-bottom: 20px;
  min-width: 500px;
}}
td {{
  border: {border_width}px solid {border_color};
  padding: 4px 8px;
}}
</style>
</head>
<body>
<h1>CyNiT Certificate Decoder Export</h1>
<p>Bestand: {info.get("filename", "")}</p>
<p>Type: {info.get("type", "")}</p>
{subject_html}
{issuer_html}
{props_html}
</body>
</html>
"""
    return html


def build_markdown_export(info: Dict[str, Any], settings: Dict[str, Any]) -> str:
    styles = load_export_styles(settings)
    mcfg = styles.get("md", {})

    title_prefix = mcfg.get("title_prefix", "# ")
    section_prefix = mcfg.get("section_prefix", "## ")
    bold_field = mcfg.get("bold_field_names", True)

    def dict_to_md_table(title_txt: str, mapping: Optional[Dict[str, Any]], issuer: bool = False) -> str:
        if issuer and mapping is None:
            return f"{section_prefix}{title_txt}\n\nCSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat.\n"
        if mapping is None:
            return ""
        lines = [f"{section_prefix}{title_txt}", "", "| Field | Value |", "| --- | --- |"]
        for k, v in mapping.items():
            field = f"**{k}**" if bold_field else k
            lines.append(f"| {field} | {v} |")
        lines.append("")
        return "\n".join(lines)

    md_parts: List[str] = []
    md_parts.append(f"{title_prefix}CyNiT Certificate Decoder Export\n")
    md_parts.append(f"**Bestand**: `{info.get('filename', '')}`  ")
    md_parts.append(f"**Type**: {info.get('type', '')}\n")

    md_parts.append(dict_to_md_table("Certificate Subject", info.get("subject")))
    md_parts.append(dict_to_md_table("Certificate Issuer", info.get("issuer"), issuer=True))
    md_parts.append(dict_to_md_table("Certificate Properties", info.get("properties")))

    return "\n".join(md_parts)


def build_xlsx_export(info: Dict[str, Any], settings: Dict[str, Any]) -> bytes:
    styles = load_export_styles(settings)
    xcfg = styles.get("xlsx", {})

    sheet_cfg = xcfg.get("sheet", {})
    title_cfg = xcfg.get("title", {})
    field_cfg = xcfg.get("field_col", {})
    value_cfg = xcfg.get("value_col", {})

    default_bg = sheet_cfg.get("default_bg", "#FFFFFF")

    title_font_size = title_cfg.get("font_size", 16)
    title_bold = title_cfg.get("bold", True)
    title_italic = title_cfg.get("italic", False)
    title_color = title_cfg.get("color", "#000000").lstrip("#")

    field_font_size = field_cfg.get("font_size", 12)
    field_bold = field_cfg.get("bold", True)
    field_italic = field_cfg.get("italic", True)
    field_color = field_cfg.get("color", "#000000").lstrip("#")

    value_font_size = value_cfg.get("font_size", 12)
    value_bold = value_cfg.get("bold", False)
    value_italic = value_cfg.get("italic", False)
    value_color = value_cfg.get("color", "#000000").lstrip("#")

    default_bg_hex = default_bg.lstrip("#")

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate"

    title_font = Font(
        color=title_color,
        bold=title_bold,
        italic=title_italic,
        size=title_font_size,
    )
    field_font = Font(
        color=field_color,
        bold=field_bold,
        italic=field_italic,
        size=field_font_size,
    )
    value_font = Font(
        color=value_color,
        bold=value_bold,
        italic=value_italic,
        size=value_font_size,
    )

    fill_default = PatternFill(
        start_color=default_bg_hex,
        end_color=default_bg_hex,
        fill_type="solid",
    )

    row = 1
    ws["A1"] = "CyNiT Certificate Decoder Export"
    ws["A1"].font = title_font
    ws["A1"].fill = fill_default
    row += 2

    def set_row(key: str, value: str):
        nonlocal row
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = field_font
        ws[f"B{row}"].font = value_font
        ws[f"A{row}"].fill = fill_default
        ws[f"B{row}"].fill = fill_default
        row += 1

    set_row("Bestand", str(info.get("filename", "")))
    set_row("Type", str(info.get("type", "")))
    row += 1

    def write_section(title_txt: str, mapping: Optional[Dict[str, Any]]):
        nonlocal row
        ws[f"A{row}"] = title_txt
        ws[f"A{row}"].font = title_font
        ws[f"A{row}"].fill = fill_default
        row += 1
        if mapping is None:
            ws[f"A{row}"] = "CSR heeft geen issuer; dit wordt pas ingevuld na uitgifte van het certificaat."
            ws[f"A{row}"].font = value_font
            ws[f"A{row}"].fill = fill_default
            row += 2
            return
        for k, v in mapping.items():
            ws[f"A{row}"] = k
            ws[f"B{row}"] = v
            ws[f"A{row}"].font = field_font
            ws[f"B{row}"].font = value_font
            ws[f"A{row}"].fill = fill_default
            ws[f"B{row}"].fill = fill_default
            row += 1
        row += 1

    write_section("Certificate Subject", info.get("subject"))
    write_section("Certificate Issuer", info.get("issuer"))
    write_section("Certificate Properties", info.get("properties"))

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_zip_bytes(info: Dict[str, Any], settings: Dict[str, Any], formats: List[str]) -> bytes:
    """
    Bouwt een ZIP met de gevraagde formaten (json, csv, xlsx, html, md).
    Returnt bytes van het zip-bestand.
    """
    base_name = Path(info.get("filename", "certificate")).stem or "certificate"

    buf = BytesIO()
    with ZipFile(buf, "w") as z:
        for fmt in formats:
            fmt = fmt.lower()
            if fmt == "json":
                content = json.dumps(info, indent=2, ensure_ascii=False).encode("utf-8")
                z.writestr(f"{base_name}.json", content)
            elif fmt == "csv":
                lines = ["Section;Field;Value"]
                for section_key, section_name in [
                    ("subject", "Subject"),
                    ("issuer", "Issuer"),
                    ("properties", "Properties"),
                ]:
                    section = info.get(section_key)
                    if section is None:
                        continue
                    for k, v in section.items():
                        value_str = str(v).replace(";", ",")
                        lines.append(f"{section_name};{k};{value_str}")
                content = "\n".join(lines).encode("utf-8")
                z.writestr(f"{base_name}.csv", content)
            elif fmt == "html":
                html = build_html_export(info, settings)
                z.writestr(f"{base_name}.html", html.encode("utf-8"))
            elif fmt == "md":
                md = build_markdown_export(info, settings)
                z.writestr(f"{base_name}.md", md.encode("utf-8"))
            elif fmt == "xlsx":
                xlsx_bytes = build_xlsx_export(info, settings)
                z.writestr(f"{base_name}.xlsx", xlsx_bytes)
            else:
                continue

    buf.seek(0)
    return buf.getvalue()
